"""
Roster auto-scheduler service.

Python port of the JavaScript scheduling engine from roster-suite.
Implements the same deterministic rotation algorithm as computeShufflePlanForMonth
and autoFixCoverageForVL from roster-suite/public/index.html.

Phase 1: standalone service — no Flask routes yet.
Phase 2: routes/roster_scheduler.py Blueprint will call these functions.
"""
import json
import logging
from datetime import date

from models.models import db, TeamMember, ShiftRoster
from models.roster_scheduler_models import (
    ScheduledShift,
    ShiftCoverageRequirement,
    PublicHoliday,
)
from services.roster_scheduler_algo import (
    DEFAULT_SHIFT_MAP as _DEFAULT_SHIFT_MAP,
    PROTECTED_CODES as _PROTECTED_CODES,
    LEAVE_CODES as _LEAVE_CODES,
    days_in_month as _days_in_month,
    is_weekend as _is_weekend,
    compute_month_plan,
)

logger = logging.getLogger(__name__)


def _get_shift_map(account_id: int) -> dict[str, str]:
    """Return shift code map for account, falling back to defaults."""
    try:
        from models.app_config import AppConfig
        raw = AppConfig.get_value('roster_shift_code_map', account_id=account_id)
        if raw:
            return {**_DEFAULT_SHIFT_MAP, **json.loads(raw)}
    except Exception:
        pass
    return _DEFAULT_SHIFT_MAP


# ---------------------------------------------------------------------------
# Core: generate_month_schedule
# ---------------------------------------------------------------------------

def generate_month_schedule(
    team_id: int,
    account_id: int,
    year: int,
    month: int,
    overwrite: bool = False,
    dry_run: bool = False,
    created_by_id: int | None = None,
) -> dict:
    """
    Generate shift assignments for all active TeamMembers of team_id for year/month.

    Returns:
        {
          "assignments": [{"team_member_id": int, "name": str,
                           "shift_date": str, "shift_code": str}, ...],
          "warnings": [...],
          "dry_run": bool,
        }

    When dry_run=False and overwrite=False: only fills days with no existing assignment.
    When dry_run=False and overwrite=True : overwrites non-protected rows.
    When dry_run=True: nothing is written.
    """
    reqs_raw = ShiftCoverageRequirement.get_team_requirements(team_id)
    if not reqs_raw:
        reqs_raw = ShiftCoverageRequirement.default_requirements()

    holidays = PublicHoliday.get_for_month(account_id, year, month)
    existing_grid = ScheduledShift.get_month_grid(team_id, year, month)

    db_members = (
        TeamMember.query
        .filter_by(team_id=team_id, account_id=account_id, is_active=True)
        .order_by(TeamMember.id)
        .all()
    )
    if not db_members:
        return {"assignments": [], "warnings": ["No active team members found. Add members to this team before generating a schedule."], "dry_run": dry_run}

    members = [
        {
            "id": m.id,
            "name": m.name,
            "scheduling_role": getattr(m, 'scheduling_role', 'support') or 'support',
            "lead_shift": getattr(m, 'lead_shift', None) or 'E',
        }
        for m in db_members
        if (getattr(m, 'scheduling_role', 'support') or 'support') != 'excluded'
    ]

    support_count = sum(1 for m in members if m["scheduling_role"] != "lead")
    if support_count == 0:
        return {"assignments": [], "warnings": [
            f"No support members found — only {len(members)} lead(s) in this team. "
            "Assign at least one member as 'support' in Scheduler Settings before generating."
        ], "dry_run": dry_run}

    plan = compute_month_plan(
        year=year,
        month=month,
        members=members,
        reqs=reqs_raw,
        holidays=holidays,
        existing=existing_grid,
        team_index=0,
    )

    if not plan:
        return {"assignments": [], "warnings": ["No support members to schedule."], "dry_run": dry_run}

    assignments = [
        {
            "team_member_id": a["member_id"],
            "name": a["name"],
            "shift_date": a["shift_date"].isoformat(),
            "shift_code": a["shift_code"],
        }
        for a in plan
    ]

    if not dry_run:
        for a in plan:
            ScheduledShift.upsert(
                team_member_id=a["member_id"],
                team_id=team_id,
                account_id=account_id,
                shift_date=a["shift_date"],
                shift_code=a["shift_code"],
                source='auto',
                created_by_id=created_by_id,
            )
        db.session.commit()

    return {"assignments": assignments, "warnings": [], "dry_run": dry_run}


def preview_month_schedule(
    team_id: int, account_id: int, year: int, month: int
) -> dict:
    """Dry-run wrapper — returns schedule without writing."""
    return generate_month_schedule(team_id, account_id, year, month, dry_run=True)


# ---------------------------------------------------------------------------
# Coverage auto-fix (mirrors autoFixCoverageForVL)
# ---------------------------------------------------------------------------

def _auto_fix_coverage(
    team_member_id: int,
    day: date,
    team_id: int,
    account_id: int,
) -> list[dict]:
    """
    After a leave code is applied to team_member_id on day, check coverage
    requirements and reassign OFF/unassigned colleagues if a shift falls short.

    Returns list of {name, shift_code, shift_date} for each fix applied.
    """
    reqs_raw = ShiftCoverageRequirement.get_team_requirements(team_id)
    if not reqs_raw:
        reqs_raw = ShiftCoverageRequirement.default_requirements()

    all_members = (
        TeamMember.query
        .filter_by(team_id=team_id, account_id=account_id, is_active=True)
        .order_by(TeamMember.id)
        .all()
    )

    existing: dict[int, str] = {}
    for m in all_members:
        row = ScheduledShift.query.filter_by(
            team_member_id=m.id, shift_date=day
        ).first()
        existing[m.id] = row.shift_code if row else ''

    fixes: list[dict] = []
    is_wknd = _is_weekend(day)

    if is_wknd:
        needed_shift = 'WMO' if day.weekday() == 5 else 'WEO'
        req = reqs_raw.get(needed_shift, 1)
        if req == '*' or req == 0:
            return fixes
        req = int(req)
        current_count = sum(
            1 for m in all_members
            if m.id != team_member_id and existing.get(m.id) == needed_shift
        )
        if current_count < req:
            shortage = req - current_count
            candidates = [
                m for m in all_members
                if m.id != team_member_id and (not existing.get(m.id) or existing.get(m.id) == 'WEO')
            ]
            if len(candidates) < shortage:
                opposite = 'WEO' if needed_shift == 'WMO' else 'WMO'
                for m in all_members:
                    if m.id == team_member_id:
                        continue
                    if any(c.id == m.id for c in candidates):
                        continue
                    if existing.get(m.id) == opposite:
                        candidates.append(m)
            for m in candidates[:shortage]:
                ScheduledShift.upsert(
                    team_member_id=m.id, team_id=team_id, account_id=account_id,
                    shift_date=day, shift_code=needed_shift, source='auto'
                )
                fixes.append({"name": m.name, "shift_code": needed_shift,
                               "shift_date": day.isoformat()})
    else:
        fixed_ids: set[int] = set()
        for shift_key in ('D', 'E'):
            req = reqs_raw.get(shift_key, 0)
            if req == '*' or not req or req == 0:
                continue
            req = int(req)
            current_count = sum(
                1 for m in all_members
                if m.id != team_member_id
                and (existing.get(m.id) == shift_key
                     or (shift_key == 'E' and existing.get(m.id) == 'N'))
            )
            if current_count < req:
                shortage = req - current_count
                candidates = [
                    m for m in all_members
                    if m.id != team_member_id
                    and m.id not in fixed_ids
                    and (not existing.get(m.id) or existing.get(m.id) == 'WEO')
                ]
                for m in candidates[:shortage]:
                    ScheduledShift.upsert(
                        team_member_id=m.id, team_id=team_id, account_id=account_id,
                        shift_date=day, shift_code=shift_key, source='auto'
                    )
                    fixed_ids.add(m.id)
                    fixes.append({"name": m.name, "shift_code": shift_key,
                                  "shift_date": day.isoformat()})

    db.session.commit()
    return fixes


# ---------------------------------------------------------------------------
# Leave application
# ---------------------------------------------------------------------------

def apply_leave(
    team_member_id: int,
    dates: list[date],
    leave_code: str,
    account_id: int,
    team_id: int,
    auto_fix: bool = True,
) -> dict:
    """
    Mark leave_code on the given dates (protected=True) and optionally
    trigger _auto_fix_coverage for each affected day.

    leave_code must be one of: VL, SL, HL, CO
    Returns: {"applied": [...], "coverage_fixes": [...]}
    """
    if leave_code not in _LEAVE_CODES:
        raise ValueError(f"leave_code must be one of {sorted(_LEAVE_CODES)}, got {leave_code!r}")

    applied: list[dict] = []
    coverage_fixes: list[dict] = []

    for d in dates:
        ScheduledShift.upsert(
            team_member_id=team_member_id,
            team_id=team_id,
            account_id=account_id,
            shift_date=d,
            shift_code=leave_code,
            source='manual',
        )
        applied.append({"shift_date": d.isoformat(), "shift_code": leave_code})

        if auto_fix:
            fixes = _auto_fix_coverage(team_member_id, d, team_id, account_id)
            coverage_fixes.extend(fixes)

    db.session.commit()
    return {"applied": applied, "coverage_fixes": coverage_fixes}


# ---------------------------------------------------------------------------
# Manual cell update
# ---------------------------------------------------------------------------

_VALID_SHIFT_CODES = {
    'D', 'E', 'N', 'OCN',
    'D/OCN', 'E/OCN', 'N/OCN',  # combined on-call night
    'WMO', 'WEO', 'OS', 'OF',
    'VL', 'SL', 'HL', 'CO', 'LE', 'G', '',
}


def update_shift(
    team_member_id: int,
    shift_date: date,
    shift_code: str,
    team_id: int,
    account_id: int,
    created_by_id: int | None = None,
) -> dict:
    """
    Manually update (or clear) a single shift cell.

    shift_code='' deletes the row.  Leave codes are auto-marked protected.
    Admin-initiated so source='manual', which bypasses the auto-only guard.
    """
    if shift_code not in _VALID_SHIFT_CODES:
        raise ValueError(f"Invalid shift_code {shift_code!r}. Allowed: {sorted(_VALID_SHIFT_CODES)}")

    if not shift_code:
        row = ScheduledShift.query.filter_by(
            team_member_id=team_member_id, shift_date=shift_date
        ).first()
        if row:
            db.session.delete(row)
    else:
        ScheduledShift.upsert(
            team_member_id=team_member_id,
            team_id=team_id,
            account_id=account_id,
            shift_date=shift_date,
            shift_code=shift_code,
            source='manual',
            created_by_id=created_by_id,
        )
    db.session.commit()
    return {
        "team_member_id": team_member_id,
        "shift_date": shift_date.isoformat(),
        "shift_code": shift_code,
    }


# ---------------------------------------------------------------------------
# Bulk fill
# ---------------------------------------------------------------------------

def bulk_fill(
    team_id: int,
    account_id: int,
    shift_code: str,
    dates: list[date],
    member_ids: list[int] | None = None,
    skip_protected: bool = True,
    created_by_id: int | None = None,
) -> dict:
    """
    Apply shift_code to every member (or a subset) across the given dates.

    skip_protected=True (default): rows with is_protected=True are left untouched.
    """
    if shift_code not in _VALID_SHIFT_CODES or not shift_code:
        raise ValueError(f"Invalid shift_code {shift_code!r}")

    members = (
        TeamMember.query
        .filter_by(team_id=team_id, account_id=account_id, is_active=True)
        .all()
    )
    if member_ids:
        id_set = set(member_ids)
        members = [m for m in members if m.id in id_set]

    if not members or not dates:
        return {"applied_count": 0}

    # Batch-fetch protected rows for the date range to avoid N*M queries
    protected_set: set[tuple] = set()
    if skip_protected:
        rows = ScheduledShift.query.filter(
            ScheduledShift.team_id == team_id,
            ScheduledShift.shift_date.in_(dates),
            ScheduledShift.is_protected == True,
        ).all()
        protected_set = {(r.team_member_id, r.shift_date) for r in rows}

    applied_count = 0
    for d in dates:
        for m in members:
            if (m.id, d) in protected_set:
                continue
            ScheduledShift.upsert(
                team_member_id=m.id,
                team_id=team_id,
                account_id=account_id,
                shift_date=d,
                shift_code=shift_code,
                source='manual',
                created_by_id=created_by_id,
            )
            applied_count += 1

    db.session.commit()
    return {"applied_count": applied_count}


# ---------------------------------------------------------------------------
# Reset month
# ---------------------------------------------------------------------------

def reset_month_schedule(team_id: int, account_id: int, year: int, month: int) -> dict:
    """
    Delete all non-protected ScheduledShift rows for team/month.

    Protected rows (leave codes) are preserved.
    Returns {"deleted_count": N}.
    """
    import calendar as _calendar
    start = date(year, month, 1)
    end = date(year, month, _calendar.monthrange(year, month)[1])

    rows = ScheduledShift.query.filter(
        ScheduledShift.team_id == team_id,
        ScheduledShift.shift_date >= start,
        ScheduledShift.shift_date <= end,
        ScheduledShift.is_protected == False,
    ).all()

    count = len(rows)
    for r in rows:
        db.session.delete(r)
    db.session.commit()
    return {"deleted_count": count}


# ---------------------------------------------------------------------------
# Export to XLSX
# ---------------------------------------------------------------------------

def export_to_xlsx(data: list[dict], year: int, month: int) -> bytes:
    """
    Build an Excel workbook from the shift view data.

    data: list of {team_member_id, name, scheduling_role, shifts}
    Returns raw bytes of the .xlsx file.
    """
    import calendar as _cal
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f'{_cal.month_name[month]} {year}'

    # Colour map for shift codes
    CODE_FILLS = {
        'D':   PatternFill('solid', fgColor='1A73E8'),
        'E':   PatternFill('solid', fgColor='F57C00'),
        'N':   PatternFill('solid', fgColor='7B1FA2'),
        'OCN': PatternFill('solid', fgColor='006064'),
        'WMO': PatternFill('solid', fgColor='F0A830'),
        'WEO': PatternFill('solid', fgColor='FF7C98'),
        'OS':  PatternFill('solid', fgColor='2E7D32'),
        'OF':  PatternFill('solid', fgColor='00838F'),
        'VL':  PatternFill('solid', fgColor='C62828'),
        'SL':  PatternFill('solid', fgColor='AD1457'),
        'HL':  PatternFill('solid', fgColor='6A1B9A'),
        'CO':  PatternFill('solid', fgColor='4E342E'),
    }
    WHITE_FONT   = Font(bold=True, color='FFFFFF', size=8)
    HEADER_FILL  = PatternFill('solid', fgColor='E8F0FE')
    HEADER_FONT  = Font(bold=True, size=9)
    WKND_FILL    = PatternFill('solid', fgColor='F5F5F5')
    CENTER       = Alignment(horizontal='center', vertical='center')
    thin         = Side(style='thin', color='D0D0D0')
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

    days = _days_in_month(year, month)
    DAYS_SHORT = ['Su', 'Mo', 'Tu', 'We', 'Th', 'Fr', 'Sa']

    # Header row: Member | 1-Su | 2-Mo | ...
    ws.cell(1, 1, 'Member').font     = HEADER_FONT
    ws.cell(1, 1).fill               = HEADER_FILL
    ws.cell(1, 1).alignment          = CENTER
    ws.column_dimensions['A'].width  = 22

    for d in range(1, days + 1):
        from datetime import date as _date
        dw = _date(year, month, d).weekday()      # Mon=0, Sun=6
        dow = DAYS_SHORT[(dw + 1) % 7]           # shift to Sun=0
        col = d + 1
        cell = ws.cell(1, col, f'{d}\n{dow}')
        cell.font      = HEADER_FONT
        cell.fill      = WKND_FILL if dw >= 5 else HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 5

    # Data rows
    for row_idx, member in enumerate(data, start=2):
        name = member['name']
        if member.get('scheduling_role') == 'lead':
            name = f'[L] {name}'
        ws.cell(row_idx, 1, name).font      = Font(bold=True, size=9)
        ws.cell(row_idx, 1).alignment       = Alignment(vertical='center')
        ws.cell(row_idx, 1).border         = BORDER

        for d in range(1, days + 1):
            from datetime import date as _date
            ds   = _date(year, month, d).isoformat()
            code = (member.get('shifts') or {}).get(ds, '')
            col  = d + 1
            cell = ws.cell(row_idx, col, code)
            cell.alignment = CENTER
            cell.border    = BORDER
            cell.font      = Font(size=8)
            if code in CODE_FILLS:
                cell.fill = CODE_FILLS[code]
                cell.font = WHITE_FONT

            dw = _date(year, month, d).weekday()
            if not code and dw >= 5:
                cell.fill = WKND_FILL

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = 'B2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Coverage view
# ---------------------------------------------------------------------------

def get_coverage_view(team_id: int, account_id: int, year: int, month: int) -> dict:
    """
    For each day of the month, count how many members are on each shift code.
    Returns:
    {
      "days": [1, 2, ...],
      "requirements": {"D": 1, "E": "*", "N": 1, ...},
      "coverage": {
        "D":  [1, 2, 1, ...],   # count per day
        "E":  [...],
        ...
      },
      "status": {               # "ok" | "short" | "empty" per shift per day
        "D":  ["ok", "short", ...],
        ...
      }
    }
    """
    import calendar as _cal
    from datetime import date as _date

    total_days  = _days_in_month(year, month)
    grid        = ScheduledShift.get_month_grid(team_id, year, month)
    reqs_raw    = ShiftCoverageRequirement.get_team_requirements(team_id)
    if not reqs_raw:
        reqs_raw = ShiftCoverageRequirement.default_requirements()

    SHIFT_CODES = ['D', 'E', 'N', 'OCN', 'WMO', 'WEO', 'OS', 'OF']

    counts: dict[str, list[int]] = {code: [0] * total_days for code in SHIFT_CODES}
    status: dict[str, list[str]] = {code: ['empty'] * total_days for code in SHIFT_CODES}

    # Tally per day
    for member_id, day_map in grid.items():
        for day, code in day_map.items():
            if code in SHIFT_CODES:
                counts[code][day - 1] += 1

    # Evaluate against requirements
    for code in SHIFT_CODES:
        req = reqs_raw.get(code, 0)
        try:
            req_int = int(req) if req != '*' else 0
        except (ValueError, TypeError):
            req_int = 0
        for i in range(total_days):
            actual = counts[code][i]
            if actual == 0:
                status[code][i] = 'empty'
            elif req == '*' or actual >= req_int:
                status[code][i] = 'ok'
            else:
                status[code][i] = 'short'

    # Serialisable requirements
    req_display = {code: str(reqs_raw.get(code, 0)) for code in SHIFT_CODES}

    return {
        'days':         list(range(1, total_days + 1)),
        'requirements': req_display,
        'coverage':     counts,
        'status':       status,
    }


# ---------------------------------------------------------------------------
# Member summary
# ---------------------------------------------------------------------------

def get_member_summary(team_id: int, account_id: int, year: int, month: int) -> list[dict]:
    """
    Per-member shift code breakdown for the month.

    Returns:
    [
      {
        "team_member_id": int,
        "name": str,
        "scheduling_role": str,
        "counts": {"D": 2, "E": 5, "N": 3, "OS": 2, "OF": 2, "VL": 0, ...},
        "total_days": 14,      # days with any assignment
      }
    ]
    """
    ALL_CODES = ['D', 'E', 'N', 'OCN', 'WMO', 'WEO', 'OS', 'OF', 'VL', 'SL', 'HL', 'CO']

    members = (
        TeamMember.query
        .filter_by(team_id=team_id, account_id=account_id, is_active=True)
        .order_by(TeamMember.id)
        .all()
    )
    grid = ScheduledShift.get_month_grid(team_id, year, month)

    result = []
    for m in members:
        day_map = grid.get(m.id, {})
        counts  = {code: 0 for code in ALL_CODES}
        for code in day_map.values():
            if code in counts:
                counts[code] += 1
        result.append({
            'team_member_id': m.id,
            'name':           m.name,
            'scheduling_role': getattr(m, 'scheduling_role', 'support'),
            'counts':         counts,
            'total_days':     len(day_map),
        })
    return result


# ---------------------------------------------------------------------------
# CSV import
# ---------------------------------------------------------------------------

_IMPORT_CODES = {'D', 'E', 'N', 'OCN', 'D/OCN', 'E/OCN', 'N/OCN', 'WMO', 'WEO', 'OS', 'OF', 'VL', 'SL', 'HL', 'CO', 'LE', 'G'}


def import_from_csv(stream, team_id: int, account_id: int, created_by_id: int | None = None) -> dict:
    """
    Import shifts from a CSV file stream.

    Expected CSV columns (case-insensitive, order flexible):
        member_name  OR  member_id   — identifies the team member
        date                         — ISO date YYYY-MM-DD
        shift_code                   — one of the valid shift codes

    Returns:
        {"imported": N, "skipped": N, "errors": [...]}
    """
    import csv
    import io

    members = (
        TeamMember.query
        .filter_by(team_id=team_id, account_id=account_id, is_active=True)
        .all()
    )
    name_map: dict[str, int] = {m.name.lower().strip(): m.id for m in members}
    id_set:   set[int]       = {m.id for m in members}

    text = stream.read()
    if isinstance(text, bytes):
        text = text.decode('utf-8-sig')       # handle BOM
    reader = csv.DictReader(io.StringIO(text))

    # Normalise column names
    if reader.fieldnames is None:
        raise ValueError('CSV appears to be empty or has no header row.')
    headers = [h.lower().strip() for h in reader.fieldnames]

    has_name = 'member_name' in headers
    has_id   = 'member_id'   in headers
    if not has_name and not has_id:
        raise ValueError("CSV must have a 'member_name' or 'member_id' column.")
    if 'date' not in headers:
        raise ValueError("CSV must have a 'date' column (YYYY-MM-DD).")
    if 'shift_code' not in headers:
        raise ValueError("CSV must have a 'shift_code' column.")

    imported = 0
    skipped  = 0
    errors:  list[str] = []

    for row_num, raw_row in enumerate(reader, start=2):
        row = {k.lower().strip(): (v or '').strip() for k, v in raw_row.items()}

        # Resolve member
        member_id = None
        if has_id and row.get('member_id'):
            try:
                mid = int(row['member_id'])
                if mid in id_set:
                    member_id = mid
            except ValueError:
                pass
        if member_id is None and has_name and row.get('member_name'):
            member_id = name_map.get(row['member_name'].lower().strip())

        if member_id is None:
            errors.append(f'Row {row_num}: member not found — "{row.get("member_name") or row.get("member_id")}"')
            skipped += 1
            continue

        # Parse date
        try:
            from datetime import date as _date
            shift_date = _date.fromisoformat(row['date'])
        except ValueError:
            errors.append(f'Row {row_num}: invalid date "{row["date"]}"')
            skipped += 1
            continue

        # Validate code
        shift_code = row['shift_code'].upper()
        if shift_code not in _IMPORT_CODES:
            errors.append(f'Row {row_num}: unknown shift code "{shift_code}"')
            skipped += 1
            continue

        ScheduledShift.upsert(
            team_member_id=member_id,
            team_id=team_id,
            account_id=account_id,
            shift_date=shift_date,
            shift_code=shift_code,
            source='import',
            created_by_id=created_by_id,
        )
        imported += 1

    db.session.commit()
    return {'imported': imported, 'skipped': skipped, 'errors': errors[:20]}


# ---------------------------------------------------------------------------
# Leave planner Excel upload
# ---------------------------------------------------------------------------

def upload_leave_plan(file_stream, team_id: int, account_id: int, leave_type: str = 'VL', created_by_id: int | None = None) -> dict:
    """
    Parse a leave planner Excel file and create leave entries.

    Expected format:
      Row 0 (header): "Resource Name" | "Stream" | "Jan-26" | "Feb-26" | ...
      Data rows:       employee name  | stream   | "12,13,14" | "1,2,25" | ...

    Day values are comma-separated day numbers. Ranges like "1-5" are also accepted.
    """
    import openpyxl
    from io import BytesIO

    _MONTH_ABBR = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                   'jul', 'aug', 'sep', 'oct', 'nov', 'dec']

    raw = file_stream.read()
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('File is empty.')

    headers = [str(h).strip() if h is not None else '' for h in rows[0]]

    # Build col_index → (year, month) for date columns
    month_cols: dict[int, tuple[int, int]] = {}
    for i, h in enumerate(headers):
        parts = h.lower().replace(' ', '-').split('-')
        if len(parts) == 2 and parts[0] in _MONTH_ABBR:
            mon = _MONTH_ABBR.index(parts[0]) + 1
            try:
                yr_raw = parts[1]
                yr = int('20' + yr_raw) if len(yr_raw) == 2 else int(yr_raw)
                month_cols[i] = (yr, mon)
            except ValueError:
                pass

    if not month_cols:
        raise ValueError('No month columns found. Expected headers like "Jan-26".')

    # Load team members
    members = (
        TeamMember.query
        .filter_by(team_id=team_id, account_id=account_id, is_active=True)
        .all()
    )
    name_map: dict[str, 'TeamMember'] = {m.name.lower().strip(): m for m in members}

    def _resolve_member(raw_name: str):
        key = raw_name.lower().strip()
        # 1. Exact match
        if key in name_map:
            return name_map[key]
        # 2. Unique prefix/suffix match (only if exactly one member matches)
        matches = [m for mkey, m in name_map.items()
                   if key in mkey or mkey in key]
        if len(matches) == 1:
            return matches[0]
        return None

    def _parse_days(cell_val) -> list[int]:
        if not cell_val:
            return []
        days: list[int] = []
        for part in str(cell_val).split(','):
            part = part.strip()
            if '-' in part:
                bounds = part.split('-')
                if len(bounds) == 2 and bounds[0].strip().isdigit() and bounds[1].strip().isdigit():
                    lo = max(1, int(bounds[0]))
                    hi = min(31, int(bounds[1]))
                    days.extend(range(lo, hi + 1))
            elif part.isdigit():
                day = int(part)
                if 1 <= day <= 31:
                    days.append(day)
        return days

    created = 0
    updated = 0
    skipped = 0
    unmatched: list[str] = []

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        member = _resolve_member(name)
        if not member:
            if name and name not in unmatched:
                unmatched.append(name)
            continue

        for col_idx, (yr, mon) in month_cols.items():
            cell_val = row[col_idx] if col_idx < len(row) else None
            days = _parse_days(cell_val)
            for day in days:
                try:
                    d = date(yr, mon, day)
                except ValueError:
                    continue

                existing = ScheduledShift.query.filter_by(
                    team_member_id=member.id, shift_date=d
                ).first()
                if existing and existing.shift_code in _PROTECTED_CODES:
                    skipped += 1
                    continue

                is_new = existing is None
                ScheduledShift.upsert(
                    team_member_id=member.id,
                    team_id=team_id,
                    account_id=account_id,
                    shift_date=d,
                    shift_code=leave_type,
                    source='leave_plan',
                    created_by_id=created_by_id,
                )
                if is_new:
                    created += 1
                else:
                    updated += 1

    db.session.commit()
    return {'created': created, 'updated': updated, 'skipped': skipped, 'unmatched': unmatched}


# ---------------------------------------------------------------------------
# Read: shift view
# ---------------------------------------------------------------------------

def get_shift_view(
    team_id: int,
    account_id: int,
    year: int,
    month: int,
) -> list[dict]:
    """
    Return the full roster grid for the month.

    [
      {
        "team_member_id": int,
        "name": str,
        "scheduling_role": str,
        "shifts": {"2026-05-01": "D", "2026-05-02": "E", ...}
      }
    ]
    """
    members = (
        TeamMember.query
        .filter_by(team_id=team_id, account_id=account_id, is_active=True)
        .order_by(TeamMember.id)
        .all()
    )
    grid = ScheduledShift.get_month_grid(team_id, year, month)
    total_days = _days_in_month(year, month)

    result = []
    for m in members:
        member_shifts = {}
        for day in range(1, total_days + 1):
            d = date(year, month, day)
            code = grid.get(m.id, {}).get(day, '')
            if code:
                member_shifts[d.isoformat()] = code
        result.append({
            "team_member_id": m.id,
            "name": m.name,
            "scheduling_role": getattr(m, 'scheduling_role', 'support'),
            "shifts": member_shifts,
        })
    return result


# ---------------------------------------------------------------------------
# Sync to existing ShiftRoster table
# ---------------------------------------------------------------------------

def sync_to_shift_roster(
    team_id: int,
    account_id: int,
    year: int,
    month: int,
    overwrite: bool = False,
) -> dict:
    """
    Copy the month's ScheduledShift entries into the existing ShiftRoster table
    so the auto-generated schedule appears in the main Shift Roster tab.

    overwrite=False (default): skip dates that already have a ShiftRoster entry.
    overwrite=True            : replace existing ShiftRoster entries for the month.

    Returns {"synced": N, "skipped": N, "overwritten": N}
    """
    import calendar as _cal

    start = date(year, month, 1)
    end   = date(year, month, _cal.monthrange(year, month)[1])

    # Fetch all ScheduledShift rows for this team/month
    scheduled = ScheduledShift.query.filter(
        ScheduledShift.team_id   == team_id,
        ScheduledShift.shift_date >= start,
        ScheduledShift.shift_date <= end,
    ).all()

    if not scheduled:
        return {"synced": 0, "skipped": 0, "overwritten": 0}

    # Build lookup of existing ShiftRoster rows: {(team_member_id, date): row}
    existing_rows = ShiftRoster.query.filter(
        ShiftRoster.team_id    == team_id,
        ShiftRoster.account_id == account_id,
        ShiftRoster.date       >= start,
        ShiftRoster.date       <= end,
    ).all()
    existing_map: dict[tuple, ShiftRoster] = {
        (r.team_member_id, r.date): r for r in existing_rows
    }

    synced = 0
    skipped = 0
    overwritten = 0

    for ss in scheduled:
        key = (ss.team_member_id, ss.shift_date)
        existing = existing_map.get(key)

        if existing:
            if overwrite:
                existing.shift_code = ss.shift_code
                overwritten += 1
            else:
                skipped += 1
        else:
            new_entry = ShiftRoster(
                date=ss.shift_date,
                shift_code=ss.shift_code,
                team_member_id=ss.team_member_id,
                account_id=account_id,
                team_id=team_id,
            )
            db.session.add(new_entry)
            synced += 1

    db.session.commit()
    return {"synced": synced, "skipped": skipped, "overwritten": overwritten}
