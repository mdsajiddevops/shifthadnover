"""
Handover Upload Template Generator - Enhanced Version

This module generates a simplified Excel template for bulk handover data upload.
The template is designed for quick entry with:
- Single consolidated sheet for all incidents (status-based classification)
- Separate sheets for Key Points, Change Info, and KB Updates
- Shift metadata is captured during upload, not in the template
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def generate_handover_template(output_path=None):
    """
    Generate an enhanced Excel template for handover data upload.
    
    The template focuses on ease of use with:
    - Single consolidated incidents sheet
    - Status-based classification (auto-routed during upload)
    - Separate sheets for secondary information
    
    Args:
        output_path: Path where to save the template. If None, returns the workbook object.
    
    Returns:
        If output_path is None, returns the Workbook object.
        Otherwise, saves to the specified path and returns the path.
    """
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C5AA0', end_color='2C5AA0', fill_type='solid')
    subheader_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
    example_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    required_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # ==================== Sheet 1: Incidents (Consolidated) ====================
    ws_incidents = wb.active
    ws_incidents.title = "Incidents"
    
    # Add instruction row at top
    ws_incidents.merge_cells('A1:H1')
    instruction_cell = ws_incidents['A1']
    instruction_cell.value = "INCIDENT HANDOVER - Enter all incidents below. Status determines automatic classification (Open/Active, Closed, Priority, Handover, Escalated)"
    instruction_cell.font = Font(bold=True, size=12, color='2C5AA0')
    instruction_cell.alignment = center_align
    ws_incidents.row_dimensions[1].height = 30
    
    # Headers (Row 2)
    incident_headers = [
        ("Application Name", 22, True),
        ("Incident ID", 18, True),
        ("Status", 18, True),
        ("Priority", 12, False),
        ("Assigned To", 20, False),
        ("Escalated To", 18, False),
        ("Description / Notes", 50, False),
        ("Resolution / Next Actions", 50, False)
    ]
    
    for col, (header, width, required) in enumerate(incident_headers, 1):
        cell = ws_incidents.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws_incidents.column_dimensions[get_column_letter(col)].width = width
        
        # Mark required columns
        if required:
            cell.value = f"{header} *"
    
    # Add example rows (Row 3-5) with different statuses
    examples = [
        ("OrderApp", "INC0001234", "Open", "High", "John Doe", "", "Database connection timeout during peak hours", "Monitoring - needs follow-up"),
        ("PaymentGateway", "INC0005678", "Closed", "Medium", "", "", "SSL certificate warning", "Renewed certificate, issue resolved"),
        ("InventorySystem", "INC0009999", "Escalated", "Critical", "", "Infrastructure Team", "Server capacity issue", "Escalated for server upgrade"),
    ]
    
    for row_idx, example in enumerate(examples, 3):
        for col_idx, value in enumerate(example, 1):
            cell = ws_incidents.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.font = Font(italic=True, color='666666')
            cell.alignment = left_align
    
    # Add note about examples
    ws_incidents.merge_cells('A6:H6')
    note_cell = ws_incidents['A6']
    note_cell.value = "↑ Example rows above (delete or overwrite). Enter your incidents starting from row 7 ↓"
    note_cell.font = Font(italic=True, size=10, color='999999')
    note_cell.alignment = center_align
    
    # Add empty rows for data entry (Row 7 onwards)
    for row in range(7, 57):  # 50 rows for incidents
        for col in range(1, 9):
            cell = ws_incidents.cell(row=row, column=col, value="")
            cell.border = thin_border
            if col <= 3:  # Required columns
                cell.fill = required_fill
    
    # Add Status dropdown validation
    status_validation = DataValidation(
        type="list",
        formula1='"Open,In Progress,Pending,Monitoring,Closed,Resolved,Priority,Handover,Escalated"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Please select a valid status from the dropdown"
    )
    ws_incidents.add_data_validation(status_validation)
    status_validation.add('C3:C56')
    
    # Add Priority dropdown validation
    priority_validation = DataValidation(
        type="list",
        formula1='"Low,Medium,High,Critical,Emergency"',
        allow_blank=True
    )
    ws_incidents.add_data_validation(priority_validation)
    priority_validation.add('D3:D56')
    
    # ==================== Sheet 2: Key Points ====================
    ws_keypoints = wb.create_sheet("Key Points")
    
    # Instruction row
    ws_keypoints.merge_cells('A1:D1')
    ws_keypoints['A1'].value = "KEY POINTS - Important items for next shift attention"
    ws_keypoints['A1'].font = Font(bold=True, size=12, color='2C5AA0')
    ws_keypoints['A1'].alignment = center_align
    ws_keypoints.row_dimensions[1].height = 30
    
    kp_headers = [
        ("Key Point Details", 60, True),
        ("Assigned To", 25, False),
        ("Status", 15, False),
        ("Priority", 12, False)
    ]
    
    for col, (header, width, required) in enumerate(kp_headers, 1):
        cell = ws_keypoints.cell(row=2, column=col, value=f"{header} *" if required else header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws_keypoints.column_dimensions[get_column_letter(col)].width = width
    
    # Example row
    example_kp = ("Monitor server CPU usage - spikes observed at peak hours. Check at 2 PM.", "David Lee", "Open", "High")
    for col, value in enumerate(example_kp, 1):
        cell = ws_keypoints.cell(row=3, column=col, value=value)
        cell.fill = example_fill
        cell.border = thin_border
        cell.font = Font(italic=True, color='666666')
    
    # Note row
    ws_keypoints.merge_cells('A4:D4')
    ws_keypoints['A4'].value = "↑ Example row above. Enter your key points starting from row 5 ↓"
    ws_keypoints['A4'].font = Font(italic=True, size=10, color='999999')
    ws_keypoints['A4'].alignment = center_align
    
    # Empty rows
    for row in range(5, 25):
        for col in range(1, 5):
            cell = ws_keypoints.cell(row=row, column=col, value="")
            cell.border = thin_border
            if col == 1:
                cell.fill = required_fill
    
    # Status dropdown for key points
    kp_status_validation = DataValidation(
        type="list",
        formula1='"Open,In Progress,Closed"',
        allow_blank=True
    )
    ws_keypoints.add_data_validation(kp_status_validation)
    kp_status_validation.add('C3:C24')
    
    # Priority dropdown
    kp_priority_validation = DataValidation(
        type="list",
        formula1='"Low,Medium,High"',
        allow_blank=True
    )
    ws_keypoints.add_data_validation(kp_priority_validation)
    kp_priority_validation.add('D3:D24')
    
    # ==================== Sheet 3: Change Information ====================
    ws_changes = wb.create_sheet("Change Info")
    
    ws_changes.merge_cells('A1:F1')
    ws_changes['A1'].value = "CHANGE INFORMATION - Scheduled or completed changes during this shift"
    ws_changes['A1'].font = Font(bold=True, size=12, color='2C5AA0')
    ws_changes['A1'].alignment = center_align
    ws_changes.row_dimensions[1].height = 30
    
    change_headers = [
        ("Application Name", 22, True),
        ("Change Number", 18, True),
        ("Description", 45, False),
        ("Scheduled Date/Time", 22, False),
        ("Status", 15, False),
        ("Notes", 40, False)
    ]
    
    for col, (header, width, required) in enumerate(change_headers, 1):
        cell = ws_changes.cell(row=2, column=col, value=f"{header} *" if required else header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws_changes.column_dimensions[get_column_letter(col)].width = width
    
    # Example row
    example_change = ("WebPortal", "CHG0001234", "SSL Certificate renewal", "2025-01-15 14:00", "Scheduled", "Coordinate with infra team")
    for col, value in enumerate(example_change, 1):
        cell = ws_changes.cell(row=3, column=col, value=value)
        cell.fill = example_fill
        cell.border = thin_border
        cell.font = Font(italic=True, color='666666')
    
    ws_changes.merge_cells('A4:F4')
    ws_changes['A4'].value = "↑ Example row above. Enter change info starting from row 5 ↓"
    ws_changes['A4'].font = Font(italic=True, size=10, color='999999')
    ws_changes['A4'].alignment = center_align
    
    for row in range(5, 20):
        for col in range(1, 7):
            cell = ws_changes.cell(row=row, column=col, value="")
            cell.border = thin_border
            if col <= 2:
                cell.fill = required_fill
    
    change_status_validation = DataValidation(
        type="list",
        formula1='"New,Scheduled,In Progress,Completed,Postponed,Cancelled"',
        allow_blank=True
    )
    ws_changes.add_data_validation(change_status_validation)
    change_status_validation.add('E3:E19')
    
    # ==================== Sheet 4: KB Information ====================
    ws_kb = wb.create_sheet("KB Info")
    
    ws_kb.merge_cells('A1:E1')
    ws_kb['A1'].value = "KNOWLEDGE BASE - KB articles created or updated during this shift"
    ws_kb['A1'].font = Font(bold=True, size=12, color='2C5AA0')
    ws_kb['A1'].alignment = center_align
    ws_kb.row_dimensions[1].height = 30
    
    kb_headers = [
        ("Application Name", 22, False),
        ("KB Number", 15, True),
        ("Title / Description", 50, True),
        ("Status", 15, False),
        ("Notes", 35, False)
    ]
    
    for col, (header, width, required) in enumerate(kb_headers, 1):
        cell = ws_kb.cell(row=2, column=col, value=f"{header} *" if required else header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws_kb.column_dimensions[get_column_letter(col)].width = width
    
    example_kb = ("HRPortal", "KB0001234", "Password reset procedure for SSO users", "Published", "Updated for new SSO provider")
    for col, value in enumerate(example_kb, 1):
        cell = ws_kb.cell(row=3, column=col, value=value)
        cell.fill = example_fill
        cell.border = thin_border
        cell.font = Font(italic=True, color='666666')
    
    ws_kb.merge_cells('A4:E4')
    ws_kb['A4'].value = "↑ Example row above. Enter KB info starting from row 5 ↓"
    ws_kb['A4'].font = Font(italic=True, size=10, color='999999')
    ws_kb['A4'].alignment = center_align
    
    for row in range(5, 15):
        for col in range(1, 6):
            cell = ws_kb.cell(row=row, column=col, value="")
            cell.border = thin_border
            if col in [2, 3]:
                cell.fill = required_fill
    
    kb_status_validation = DataValidation(
        type="list",
        formula1='"New,Draft,In Review,Published,Archived"',
        allow_blank=True
    )
    ws_kb.add_data_validation(kb_status_validation)
    kb_status_validation.add('D3:D14')
    
    # ==================== Sheet 5: Instructions ====================
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ("HANDOVER UPLOAD TEMPLATE - QUICK GUIDE", True, 16),
        ("", False, 11),
        ("HOW TO USE THIS TEMPLATE:", True, 12),
        ("", False, 11),
        ("1. Fill in the 'Incidents' sheet with all incident details", False, 11),
        ("   - Status determines automatic classification (Open, Closed, Priority, etc.)", False, 11),
        ("   - Required fields are marked with * and highlighted in yellow", False, 11),
        ("", False, 11),
        ("2. Optionally fill 'Key Points', 'Change Info', and 'KB Info' sheets", False, 11),
        ("", False, 11),
        ("3. Save the file and upload in the Handover Form", False, 11),
        ("", False, 11),
        ("4. On upload, you will select:", False, 11),
        ("   - Team (auto-selected if you belong to one team)", False, 11),
        ("   - Handover Date", False, 11),
        ("   - Current Shift and Next Shift", False, 11),
        ("   - Engineers will be auto-populated based on roster", False, 11),
        ("", False, 11),
        ("INCIDENT STATUS MAPPING:", True, 12),
        ("", False, 11),
        ("Status 'Open', 'In Progress', 'Pending' → Open/Active Incidents", False, 11),
        ("Status 'Closed', 'Resolved' → Closed Incidents", False, 11),
        ("Status 'Priority' → Priority Incidents (High severity)", False, 11),
        ("Status 'Handover', 'Monitoring' → Handover Incidents (needs follow-up)", False, 11),
        ("Status 'Escalated' → Escalated Incidents", False, 11),
        ("", False, 11),
        ("TIPS:", True, 12),
        ("", False, 11),
        ("- Delete the example rows before uploading", False, 11),
        ("- Use dropdowns for Status and Priority fields", False, 11),
        ("- Leave rows blank if no data for that section", False, 11),
        ("- For support, contact your team administrator", False, 11),
    ]
    
    for row_num, (text, is_bold, size) in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_num, column=1, value=text)
        cell.font = Font(bold=is_bold, size=size, color='2C5AA0' if is_bold else '333333')
    
    ws_instructions.column_dimensions['A'].width = 80
    
    # Move Instructions to be first sheet
    wb.move_sheet(ws_instructions, offset=-4)
    
    # Save or return
    if output_path:
        wb.save(output_path)
        return output_path
    return wb


def generate_template_to_static():
    """Generate the template and save it to the static folder for download."""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    
    static_dir = os.path.join(project_root, 'static', 'templates')
    os.makedirs(static_dir, exist_ok=True)
    
    output_path = os.path.join(static_dir, 'handover_upload_template.xlsx')
    generate_handover_template(output_path)
    print(f"Template generated: {output_path}")
    return output_path


if __name__ == '__main__':
    generate_template_to_static()
