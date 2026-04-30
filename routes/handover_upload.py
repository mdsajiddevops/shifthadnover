"""
Handover Upload Route - Enhanced Version

This module handles the upload and processing of Excel files containing handover data.
Features:
- Shift metadata captured during upload (not in file)
- Single consolidated incidents sheet with status-based classification
- Auto-population of engineers based on roster
- Separate sheets for Key Points, Change Info, KB Updates
"""

import os
import tempfile
from datetime import datetime, date
from flask import Blueprint, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
from openpyxl import load_workbook
from models.models import db, User, Team, TeamMember, ShiftRoster
from models.team_roster_models import RosterAssignment, TeamShiftConfig

handover_upload_bp = Blueprint('handover_upload', __name__)

# Status mapping for incident classification
STATUS_CLASSIFICATION = {
    # Open/Active Incidents
    'open': 'open',
    'in progress': 'open',
    'pending': 'open',
    'active': 'open',
    
    # Closed Incidents
    'closed': 'closed',
    'resolved': 'closed',
    'completed': 'closed',
    
    # Priority Incidents
    'priority': 'priority',
    'critical': 'priority',
    'emergency': 'priority',
    
    # Handover Incidents (need follow-up)
    'handover': 'handover',
    'monitoring': 'handover',
    'pending vendor': 'handover',
    'pending customer': 'handover',
    'waiting': 'handover',
    
    # Escalated Incidents
    'escalated': 'escalated',
}


def get_cell_value(ws, row, col):
    """Safely get cell value, returning empty string for None."""
    value = ws.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def classify_incident_status(status):
    """Classify incident based on status value."""
    if not status:
        return 'open'  # Default to open
    
    status_lower = status.lower().strip()
    return STATUS_CLASSIFICATION.get(status_lower, 'open')


def parse_uploaded_handover_enhanced(file_path):
    """
    Parse the uploaded Excel file with enhanced format.
    Single incidents sheet with status-based classification.
    
    Args:
        file_path: Path to the uploaded Excel file
        
    Returns:
        Dictionary containing parsed and classified handover data
    """
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        return {'error': f'Failed to read Excel file: {str(e)}'}
    
    result = {
        'incidents': {
            'open': [],
            'closed': [],
            'priority': [],
            'handover': [],
            'escalated': []
        },
        'key_points': [],
        'change_info': [],
        'kb_updates': [],
        'errors': [],
        'warnings': []
    }
    
    # Parse Incidents sheet (consolidated)
    if 'Incidents' in wb.sheetnames:
        ws = wb['Incidents']
        
        # Example data to skip (these are pre-filled examples in the template)
        example_apps = ['orderapp', 'paymentgateway', 'inventorysystem']
        example_incidents = ['inc0001234', 'inc0005678', 'inc0009999']
        
        # Start from row 3 (row 1 = instruction, row 2 = header)
        for row in range(3, ws.max_row + 1):
            app_name = get_cell_value(ws, row, 1)
            incident_id = get_cell_value(ws, row, 2)
            status = get_cell_value(ws, row, 3)
            
            # Skip empty rows
            if not app_name and not incident_id:
                continue
            
            # Skip example indicator rows (row with arrows)
            if '↑' in app_name or '↓' in app_name or 'example' in app_name.lower():
                continue
            
            # Skip the pre-filled example rows
            if app_name.lower().strip() in example_apps:
                continue
            if incident_id.lower().strip() in example_incidents:
                continue
            
            priority = get_cell_value(ws, row, 4) or 'Medium'
            assigned_to = get_cell_value(ws, row, 5)
            escalated_to = get_cell_value(ws, row, 6)
            description = get_cell_value(ws, row, 7)
            resolution = get_cell_value(ws, row, 8)
            
            # Classify based on status
            classification = classify_incident_status(status)
            
            incident_data = {
                'app_name': app_name,
                'incident_id': incident_id,
                'status': status,
                'priority': priority,
                'assigned_to': assigned_to,
                'escalated_to': escalated_to,
                'description': description,
                'resolution': resolution,
                'row': row
            }
            
            result['incidents'][classification].append(incident_data)
    
    # Parse Key Points sheet
    if 'Key Points' in wb.sheetnames:
        ws = wb['Key Points']
        
        # Example key point to skip
        example_keypoints = ['monitor server cpu usage']
        
        for row in range(3, ws.max_row + 1):
            details = get_cell_value(ws, row, 1)
            
            # Skip empty rows or example rows
            if not details:
                continue
            if '↑' in details or '↓' in details or 'example' in details.lower():
                continue
            
            # Skip pre-filled example
            details_lower = details.lower().strip()
            if any(ex in details_lower for ex in example_keypoints):
                continue
            
            result['key_points'].append({
                'description': details,
                'assigned_to': get_cell_value(ws, row, 2),
                'status': get_cell_value(ws, row, 3) or 'Open',
                'priority': get_cell_value(ws, row, 4) or 'Medium'
            })
    
    # Parse Change Info sheet
    if 'Change Info' in wb.sheetnames:
        ws = wb['Change Info']
        
        # Example change to skip
        example_changes = ['webportal', 'chg0001234', 'ssl certificate renewal']
        
        for row in range(3, ws.max_row + 1):
            app_name = get_cell_value(ws, row, 1)
            change_num = get_cell_value(ws, row, 2)
            
            # Skip empty rows
            if not app_name and not change_num:
                continue
            
            # Skip example indicator rows
            if '↑' in str(app_name) or '↓' in str(app_name):
                continue
            
            # Skip pre-filled examples
            if app_name.lower().strip() in example_changes:
                continue
            if change_num.lower().strip() in example_changes:
                continue
            
            result['change_info'].append({
                'app_name': app_name,
                'change_number': change_num,
                'description': get_cell_value(ws, row, 3),
                'datetime': get_cell_value(ws, row, 4),
                'status': get_cell_value(ws, row, 5) or 'New',
                'notes': get_cell_value(ws, row, 6)
            })
    
    # Parse KB Info sheet
    if 'KB Info' in wb.sheetnames:
        ws = wb['KB Info']
        
        # Example KB to skip
        example_kb = ['hrportal', 'kb0001234', 'password reset procedure']
        
        for row in range(3, ws.max_row + 1):
            app_name = get_cell_value(ws, row, 1)
            kb_num = get_cell_value(ws, row, 2)
            title = get_cell_value(ws, row, 3)
            
            # Skip empty rows
            if not kb_num and not title:
                continue
            
            # Skip example indicator rows
            if '↑' in str(kb_num) or '↓' in str(kb_num) or '↑' in str(app_name) or '↓' in str(app_name):
                continue
            
            # Skip pre-filled examples
            if app_name.lower().strip() in example_kb:
                continue
            if kb_num.lower().strip() in example_kb:
                continue
            if any(ex in title.lower() for ex in example_kb):
                continue
            
            result['kb_updates'].append({
                'app_name': app_name,
                'kb_number': kb_num,
                'description': title,
                'status': get_cell_value(ws, row, 4) or 'New',
                'notes': get_cell_value(ws, row, 5)
            })
    
    # Calculate summary
    result['summary'] = {
        'open_incidents': len(result['incidents']['open']),
        'closed_incidents': len(result['incidents']['closed']),
        'priority_incidents': len(result['incidents']['priority']),
        'handover_incidents': len(result['incidents']['handover']),
        'escalated_incidents': len(result['incidents']['escalated']),
        'total_incidents': sum(len(v) for v in result['incidents'].values()),
        'key_points': len(result['key_points']),
        'change_info': len(result['change_info']),
        'kb_updates': len(result['kb_updates']),
        'total_entries': (
            sum(len(v) for v in result['incidents'].values()) +
            len(result['key_points']) +
            len(result['change_info']) +
            len(result['kb_updates'])
        )
    }
    
    return result


@handover_upload_bp.route('/api/handover/download-template')
@login_required
def download_template():
    """Download the handover upload template."""
    from utils.handover_template_generator import generate_handover_template
    import io
    
    # Generate the template in memory
    wb = generate_handover_template()
    
    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'handover_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@handover_upload_bp.route('/api/handover/upload', methods=['POST'])
@login_required
def upload_handover():
    """
    Upload and parse a handover Excel file.
    Returns the parsed data for preview before submission.
    """
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    # Save to temporary file
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, f'handover_upload_{current_user.id}_{datetime.now().timestamp()}.xlsx')
    
    try:
        file.save(temp_path)
        
        # Parse the file with enhanced parser
        parsed_data = parse_uploaded_handover_enhanced(temp_path)
        
        # Check for errors
        if parsed_data.get('error'):
            return jsonify({
                'success': False,
                'error': parsed_data['error']
            }), 400
        
        if parsed_data.get('errors'):
            return jsonify({
                'success': False,
                'error': 'Validation errors found',
                'errors': parsed_data['errors'],
                'warnings': parsed_data.get('warnings', [])
            }), 400
        
        # Return parsed data for preview
        return jsonify({
            'success': True,
            'data': parsed_data,
            'message': f"Successfully parsed {parsed_data['summary']['total_entries']} entries from the uploaded file"
        })
        
    except Exception as e:
        current_app.logger.error(f"Error processing handover upload: {str(e)}")
        return jsonify({'success': False, 'error': f'Error processing file: {str(e)}'}), 500
    finally:
        # Clean up temp file
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass


@handover_upload_bp.route('/api/handover/get-engineers', methods=['POST'])
@login_required
def get_engineers_for_upload():
    """
    Get engineers for the specified team, date, and shifts.
    Used for auto-populating engineer lists during upload.
    """
    data = request.get_json()
    
    team_id = data.get('team_id')
    handover_date = data.get('date')
    current_shift = data.get('current_shift')
    next_shift = data.get('next_shift')
    
    if not team_id:
        return jsonify({'success': False, 'error': 'Team ID is required'}), 400
    
    # Parse date
    if handover_date:
        try:
            if isinstance(handover_date, str):
                roster_date = datetime.strptime(handover_date, '%Y-%m-%d').date()
            else:
                roster_date = handover_date
        except:
            roster_date = date.today()
    else:
        roster_date = date.today()
    
    # Shift code mapping
    shift_map = {
        'Morning': 'D',
        'Evening': 'E',
        'Late Evening': 'LE',
        'Night': 'N',
        'General': 'G',
        'OnShore': 'OS',
        'OffShore': 'OF'
    }
    
    def get_engineers_by_shift(shift_type):
        """Get engineers for a specific shift using ShiftRoster (same as main form)."""
        if not shift_type:
            return []
        
        # Get shift codes (same logic as main handover form)
        primary_code = shift_map.get(shift_type, '')
        
        # Additional shift codes to include
        additional_shift_codes = {
            'Morning': ['G', 'OS'],      # Morning also shows General and OnShore engineers
            'Evening': [],               # No additional codes for Evening
            'Late Evening': ['N'],       # Late Evening also shows Night engineers
            'Night': ['LE', 'OF'],       # Night also shows Late Evening and OffShore engineers
            'General': ['D'],            # General also shows Morning (Day) engineers
            'OnShore': ['D'],            # OnShore also shows Morning (Day) engineers
            'OffShore': ['N']            # OffShore also shows Night engineers
        }
        
        # Collect all shift codes to look for
        shift_codes = [primary_code] if primary_code else []
        shift_codes.extend(additional_shift_codes.get(shift_type, []))
        
        engineers = []
        seen_ids = set()
        
        # Query ShiftRoster for each shift code
        for code in shift_codes:
            if not code:
                continue
            
            # Query ShiftRoster (same table as main form)
            roster_entries = ShiftRoster.query.filter(
                ShiftRoster.date == roster_date,
                ShiftRoster.shift_code == code,
                ShiftRoster.team_id == team_id
            ).all()
            
            for entry in roster_entries:
                if entry.team_member_id not in seen_ids:
                    member = TeamMember.query.get(entry.team_member_id)
                    if member and member.is_active:
                        engineers.append({
                            'id': member.id,
                            'name': member.name,
                            'user_id': member.user_id,
                            'shift_name': shift_type
                        })
                        seen_ids.add(entry.team_member_id)
        
        # If no roster data found, return all team members as fallback
        if not engineers:
            team_members = TeamMember.query.filter_by(team_id=team_id, is_active=True).all()
            for member in team_members:
                engineers.append({
                    'id': member.id,
                    'name': member.name,
                    'user_id': member.user_id,
                    'shift_name': 'All'
                })
        
        return engineers
    
    current_engineers = get_engineers_by_shift(current_shift)
    next_engineers = get_engineers_by_shift(next_shift)
    
    # Get all team members for dropdown options
    all_members = TeamMember.query.filter_by(team_id=team_id, is_active=True).all()
    
    return jsonify({
        'success': True,
        'current_engineers': current_engineers,
        'next_engineers': next_engineers,
        'all_team_members': [{'id': m.id, 'name': m.name, 'user_id': m.user_id} for m in all_members]
    })


@handover_upload_bp.route('/api/handover/user-teams')
@login_required
def get_user_teams():
    """
    Get teams for the current user.
    Returns single team auto-selected if user belongs to only one team.
    """
    from models.models import Account
    
    teams = []
    
    if current_user.role == 'super_admin':
        # Super admin can see all teams
        all_teams = Team.query.filter_by(is_active=True).order_by(Team.account_id, Team.name).all()
        for team in all_teams:
            teams.append({
                'id': team.id,
                'name': team.name,
                'account_id': team.account_id,
                'account_name': team.account.name if team.account else 'Unknown'
            })
    elif current_user.role == 'account_admin':
        # Account admin can see teams in their account
        account_teams = Team.query.filter_by(
            account_id=current_user.account_id,
            is_active=True
        ).order_by(Team.name).all()
        for team in account_teams:
            teams.append({
                'id': team.id,
                'name': team.name,
                'account_id': team.account_id,
                'account_name': team.account.name if team.account else 'Unknown'
            })
    else:
        # Regular users and team admins - check team memberships
        memberships = TeamMember.query.filter_by(
            user_id=current_user.id,
            is_active=True
        ).all()
        
        for membership in memberships:
            team = Team.query.get(membership.team_id)
            if team and team.is_active:
                teams.append({
                    'id': team.id,
                    'name': team.name,
                    'account_id': team.account_id,
                    'account_name': team.account.name if team.account else 'Unknown'
                })
        
        # Fallback to assigned team if no memberships
        if not teams and current_user.team_id:
            team = Team.query.get(current_user.team_id)
            if team and team.is_active:
                teams.append({
                    'id': team.id,
                    'name': team.name,
                    'account_id': team.account_id,
                    'account_name': team.account.name if team.account else 'Unknown'
                })
    
    return jsonify({
        'success': True,
        'teams': teams,
        'auto_select': len(teams) == 1,
        'selected_team': teams[0] if len(teams) == 1 else None
    })
