"""
Roster Scheduler Blueprint — Phase 2 Flask integration.

Pages:
  GET  /roster-scheduler            — main grid view (all roles)
  GET  /admin/roster-scheduler      — admin settings (coverage reqs, holidays, member roles)

API endpoints:
  GET  /api/roster/schedule         — fetch current month grid
  POST /api/roster/generate         — generate + persist schedule
  POST /api/roster/preview          — dry-run preview
  POST /api/roster/leave            — apply leave + auto-fix
  GET  /api/roster/coverage-requirements          — list coverage reqs for team
  POST /api/roster/coverage-requirements          — create/update req
  DELETE /api/roster/coverage-requirements/<id>   — delete req
  GET  /api/roster/holidays         — list holidays for account+year
  POST /api/roster/holidays         — create holiday
  DELETE /api/roster/holidays/<id>  — delete holiday
  GET  /api/roster/members          — list team members with scheduling_role
  PUT  /api/roster/members/<id>/role — update scheduling_role
"""
from datetime import date, datetime

from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, session
from flask_login import login_required, current_user

from models.models import db, TeamMember
from models.roster_scheduler_models import (
    ScheduledShift,
    ShiftCoverageRequirement,
    PublicHoliday,
)

roster_scheduler_bp = Blueprint('roster_scheduler', __name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _admin_required():
    return current_user.role in ('super_admin', 'account_admin', 'team_admin')


def _current_account_team():
    """Return (account_id, team_id) from session or current_user."""
    from flask import session
    account_id = session.get('selected_account_id') or getattr(current_user, 'account_id', None)
    team_id = session.get('selected_team_id') or getattr(current_user, 'team_id', None)
    return account_id, team_id


def _resolve_account_for_team(team_id, fallback_account_id):
    """Return the account_id for the given team_id.
    Super admins can operate across accounts — always derive from the team
    so the API uses the correct account regardless of session state.
    """
    if not team_id:
        return fallback_account_id
    if current_user.role == 'super_admin':
        from models.models import Team
        team = Team.query.get(int(team_id))
        return team.account_id if team else fallback_account_id
    return fallback_account_id


_MIN_YEAR = 2000
_MAX_YEAR = 2100


def _clamp_year(y: int) -> int:
    return max(_MIN_YEAR, min(_MAX_YEAR, y))


def _parse_year_month(req):
    today = date.today()
    try:
        year = int(req.args.get('year', today.year))
        month = int(req.args.get('month', today.month))
    except (ValueError, TypeError):
        year, month = today.year, today.month
    year = max(_MIN_YEAR, min(_MAX_YEAR, year))
    month = max(1, min(12, month))
    return year, month


# ---------------------------------------------------------------------------
# Page views
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/roster-scheduler')
@login_required
def roster_scheduler():
    """Main roster scheduler view — available to all authenticated users."""
    account_id, team_id = _current_account_team()
    today = date.today()
    year = int(request.args.get('year', today.year))
    month = int(request.args.get('month', today.month))

    # Clamp
    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1

    # Month navigation labels
    import calendar
    month_name = calendar.month_name[month]

    from models.models import Account, Team

    # Super admin: allow account selection via query param, persist to session
    if current_user.role == 'super_admin':
        req_account_id = request.args.get('account_id', type=int)
        if req_account_id:
            account_id = req_account_id
            session['selected_account_id'] = req_account_id
        accounts = Account.query.order_by(Account.name).all()
    else:
        accounts = []

    # Fetch teams for the resolved account
    teams = Team.query.filter_by(account_id=account_id).order_by(Team.name).all() if account_id else []

    # Reset team_id if it doesn't belong to the selected account
    if team_id and not any(t.id == team_id for t in teams):
        team_id = teams[0].id if teams else None

    # Also accept team_id override from query param, persist to session
    req_team_id = request.args.get('team_id', type=int)
    if req_team_id and any(t.id == req_team_id for t in teams):
        team_id = req_team_id
        session['selected_team_id'] = req_team_id

    no_account_warning = (current_user.role == 'super_admin' and not account_id)

    return render_template(
        'roster_scheduler/index.html',
        year=year,
        month=month,
        month_name=month_name,
        accounts=accounts,
        teams=teams,
        team_id=team_id,
        account_id=account_id,
        no_account_warning=no_account_warning,
    )


@roster_scheduler_bp.route('/admin/roster-scheduler')
@login_required
def roster_scheduler_admin():
    """Admin settings page — coverage requirements, holidays, member roles."""
    if not _admin_required():
        flash('Admin access required.', 'danger')
        return redirect(url_for('roster_scheduler.roster_scheduler'))

    account_id, team_id = _current_account_team()

    from models.models import Account, Team

    # Super admin: allow account selection via query param, persist to session
    if current_user.role == 'super_admin':
        req_account_id = request.args.get('account_id', type=int)
        if req_account_id:
            account_id = req_account_id
            session['selected_account_id'] = req_account_id
        accounts = Account.query.order_by(Account.name).all()
    else:
        accounts = []

    teams = Team.query.filter_by(account_id=account_id).order_by(Team.name).all() if account_id else []
    no_account_warning = (current_user.role == 'super_admin' and not account_id)

    # Reset team_id if it doesn't belong to the selected account
    if team_id and not any(t.id == team_id for t in teams):
        team_id = teams[0].id if teams else None

    today = date.today()
    return render_template(
        'roster_scheduler/admin.html',
        accounts=accounts,
        teams=teams,
        team_id=team_id,
        account_id=account_id,
        current_year=today.year,
        no_account_warning=no_account_warning,
    )


# ---------------------------------------------------------------------------
# API — schedule
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/schedule')
@login_required
def api_get_schedule():
    """Get the published shift schedule for a team and month.
    ---
    tags:
      - roster
    security:
      - SessionCookie: []
    parameters:
      - in: query
        name: year
        type: integer
        description: Year (defaults to current year)
      - in: query
        name: month
        type: integer
        description: Month 1–12 (defaults to current month)
      - in: query
        name: team_id
        type: integer
        description: Team ID (defaults to session team)
    responses:
      200:
        description: Shift schedule grid for the requested month
        schema:
          type: object
          properties:
            success:
              type: boolean
            year:
              type: integer
            month:
              type: integer
            data:
              type: object
      400:
        description: Invalid team_id
    """
    account_id, team_id = _current_account_team()
    year, month = _parse_year_month(request)

    req_team_id = request.args.get('team_id', team_id)
    try:
        req_team_id = int(req_team_id)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid team_id'}), 400

    account_id = _resolve_account_for_team(req_team_id, account_id)

    from services.roster_scheduler_service import get_shift_view
    try:
        data = get_shift_view(req_team_id, account_id, year, month)
        return jsonify({'success': True, 'data': data, 'year': year, 'month': month})
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


@roster_scheduler_bp.route('/api/roster/generate', methods=['POST'])
@login_required
def api_generate_schedule():
    """Auto-generate a shift schedule for a team and month (admin only).
    ---
    tags:
      - roster
    security:
      - SessionCookie: []
    parameters:
      - in: body
        name: body
        required: true
        schema:
          type: object
          properties:
            team_id:
              type: integer
            year:
              type: integer
            month:
              type: integer
            overwrite:
              type: boolean
              default: false
              description: Overwrite existing schedule if present
    responses:
      200:
        description: Schedule generated successfully with synced count
      400:
        description: Invalid parameters
      403:
        description: Admin role required
    """
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()
    body = request.get_json(force=True) or {}

    try:
        req_team_id = int(body.get('team_id', team_id))
        year = _clamp_year(int(body.get('year', date.today().year)))
        month = int(body.get('month', date.today().month))
        overwrite = bool(body.get('overwrite', False))
    except (ValueError, TypeError) as exc:
        return jsonify({'success': False, 'error': f'Invalid parameters: {exc}'}), 400

    account_id = _resolve_account_for_team(req_team_id, account_id)

    from services.roster_scheduler_service import generate_month_schedule
    try:
        result = generate_month_schedule(req_team_id, account_id, year, month,
                                         overwrite=overwrite,
                                         created_by_id=current_user.id)
        return jsonify({'success': True, **result})
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


@roster_scheduler_bp.route('/api/roster/preview', methods=['POST'])
@login_required
def api_preview_schedule():
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()
    body = request.get_json(force=True) or {}

    try:
        req_team_id = int(body.get('team_id', team_id))
        year = _clamp_year(int(body.get('year', date.today().year)))
        month = int(body.get('month', date.today().month))
    except (ValueError, TypeError) as exc:
        return jsonify({'success': False, 'error': f'Invalid parameters: {exc}'}), 400

    account_id = _resolve_account_for_team(req_team_id, account_id)

    from services.roster_scheduler_service import preview_month_schedule
    try:
        result = preview_month_schedule(req_team_id, account_id, year, month)
        return jsonify({'success': True, **result})
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


@roster_scheduler_bp.route('/api/roster/leave', methods=['POST'])
@login_required
def api_apply_leave():
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()
    body = request.get_json(force=True) or {}

    try:
        member_id = int(body['member_id'])
        leave_code = str(body['leave_code'])
        raw_dates = body['dates']  # list of "YYYY-MM-DD"
        dates = [date.fromisoformat(d) for d in raw_dates]
        auto_fix = bool(body.get('auto_fix', True))
    except (KeyError, ValueError, TypeError) as exc:
        return jsonify({'success': False, 'error': f'Invalid parameters: {exc}'}), 400

    from services.roster_scheduler_service import apply_leave
    try:
        result = apply_leave(member_id, dates, leave_code, account_id, team_id, auto_fix)
        return jsonify({'success': True, **result})
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# API — coverage requirements
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/coverage-requirements', methods=['GET'])
@login_required
def api_get_coverage_reqs():
    account_id, team_id = _current_account_team()
    req_team_id = request.args.get('team_id', team_id)
    try:
        req_team_id = int(req_team_id)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid team_id'}), 400

    rows = ShiftCoverageRequirement.query.filter_by(
        team_id=req_team_id, is_active=True
    ).order_by(ShiftCoverageRequirement.shift_code).all()

    return jsonify({'success': True, 'data': [
        {'id': r.id, 'shift_code': r.shift_code, 'required_count': r.required_count}
        for r in rows
    ]})


@roster_scheduler_bp.route('/api/roster/coverage-requirements', methods=['POST'])
@login_required
def api_upsert_coverage_req():
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()
    body = request.get_json(force=True) or {}

    try:
        req_team_id = int(body.get('team_id', team_id))
        shift_code = str(body['shift_code']).strip().upper()
        required_count = str(body['required_count']).strip()
    except (KeyError, ValueError, TypeError) as exc:
        return jsonify({'success': False, 'error': f'Invalid parameters: {exc}'}), 400

    # Validate required_count
    if required_count != '*' and not required_count.isdigit():
        return jsonify({'success': False, 'error': 'required_count must be a number or *'}), 400

    existing = ShiftCoverageRequirement.query.filter_by(
        team_id=req_team_id, shift_code=shift_code
    ).first()
    if existing:
        existing.required_count = required_count
        existing.is_active = True
    else:
        existing = ShiftCoverageRequirement(
            team_id=req_team_id,
            account_id=account_id,
            shift_code=shift_code,
            required_count=required_count,
        )
        db.session.add(existing)
    db.session.commit()
    return jsonify({'success': True, 'id': existing.id})


@roster_scheduler_bp.route('/api/roster/coverage-requirements/<int:req_id>', methods=['DELETE'])
@login_required
def api_delete_coverage_req(req_id):
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    row = ShiftCoverageRequirement.query.get_or_404(req_id)
    row.is_active = False
    db.session.commit()
    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# API — public holidays
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/holidays', methods=['GET'])
@login_required
def api_get_holidays():
    account_id, _ = _current_account_team()
    try:
        year = int(request.args.get('year', date.today().year))
    except (ValueError, TypeError):
        year = date.today().year

    import calendar
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    rows = PublicHoliday.query.filter(
        PublicHoliday.account_id == account_id,
        PublicHoliday.date >= start,
        PublicHoliday.date <= end,
        PublicHoliday.is_active == True,
    ).order_by(PublicHoliday.date).all()

    return jsonify({'success': True, 'data': [
        {'id': r.id, 'date': r.date.isoformat(), 'name': r.name}
        for r in rows
    ]})


@roster_scheduler_bp.route('/api/roster/holidays', methods=['POST'])
@login_required
def api_create_holiday():
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, _ = _current_account_team()
    body = request.get_json(force=True) or {}

    try:
        holiday_date = date.fromisoformat(str(body['date']))
        name = str(body['name']).strip()
    except (KeyError, ValueError, TypeError) as exc:
        return jsonify({'success': False, 'error': f'Invalid parameters: {exc}'}), 400

    if not name:
        return jsonify({'success': False, 'error': 'Holiday name is required'}), 400

    existing = PublicHoliday.query.filter_by(
        account_id=account_id, date=holiday_date
    ).first()
    if existing:
        existing.name = name
        existing.is_active = True
    else:
        existing = PublicHoliday(account_id=account_id, date=holiday_date, name=name)
        db.session.add(existing)
    db.session.commit()
    return jsonify({'success': True, 'id': existing.id})


@roster_scheduler_bp.route('/api/roster/holidays/<int:holiday_id>', methods=['DELETE'])
@login_required
def api_delete_holiday(holiday_id):
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    row = PublicHoliday.query.get_or_404(holiday_id)
    row.is_active = False
    db.session.commit()
    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# API — team members (scheduling_role)
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/members')
@login_required
def api_get_members():
    account_id, team_id = _current_account_team()
    req_team_id = request.args.get('team_id', team_id)
    try:
        req_team_id = int(req_team_id)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid team_id'}), 400

    account_id = _resolve_account_for_team(req_team_id, account_id)

    members = TeamMember.query.filter_by(
        team_id=req_team_id, account_id=account_id, is_active=True
    ).order_by(TeamMember.name).all()

    return jsonify({'success': True, 'data': [
        {
            'id': m.id,
            'name': m.name,
            'role': m.role,
            'scheduling_role': getattr(m, 'scheduling_role', 'support') or 'support',
            'lead_shift': getattr(m, 'lead_shift', 'E') or 'E',
        }
        for m in members
    ]})


@roster_scheduler_bp.route('/api/roster/members/<int:member_id>/role', methods=['PUT'])
@login_required
def api_update_member_role(member_id):
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    body = request.get_json(force=True) or {}
    scheduling_role = str(body.get('scheduling_role', 'support')).lower()
    if scheduling_role not in ('lead', 'support', 'excluded'):
        return jsonify({'success': False, 'error': "scheduling_role must be 'lead', 'support', or 'excluded'"}), 400

    member = TeamMember.query.get_or_404(member_id)
    member.scheduling_role = scheduling_role
    if scheduling_role == 'lead' and 'lead_shift' in body:
        valid_shifts = ('D', 'E', 'N', 'OCN')
        ls = str(body['lead_shift']).upper().strip()
        if ls in valid_shifts:
            member.lead_shift = ls
    db.session.commit()
    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# API — manual cell edit
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/shift', methods=['PUT'])
@login_required
def api_update_shift():
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()
    body = request.get_json(force=True) or {}

    try:
        member_id = int(body['team_member_id'])
        shift_date = date.fromisoformat(str(body['shift_date']))
        shift_code = str(body.get('shift_code', '')).strip().upper()
        req_team_id = int(body.get('team_id', team_id))
    except (KeyError, ValueError, TypeError) as exc:
        return jsonify({'success': False, 'error': f'Invalid parameters: {exc}'}), 400

    from services.roster_scheduler_service import update_shift
    try:
        result = update_shift(member_id, shift_date, shift_code, req_team_id, account_id,
                              created_by_id=current_user.id)
        return jsonify({'success': True, **result})
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# API — bulk fill
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/bulk-fill', methods=['POST'])
@login_required
def api_bulk_fill():
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()
    body = request.get_json(force=True) or {}

    try:
        req_team_id = int(body.get('team_id', team_id))
        shift_code = str(body['shift_code']).strip().upper()
        date_from = date.fromisoformat(str(body['date_from']))
        date_to = date.fromisoformat(str(body['date_to']))
        weekdays_only = bool(body.get('weekdays_only', False))
        member_ids = body.get('member_ids')
    except (KeyError, ValueError, TypeError) as exc:
        return jsonify({'success': False, 'error': f'Invalid parameters: {exc}'}), 400

    if date_from > date_to:
        return jsonify({'success': False, 'error': 'date_from must be <= date_to'}), 400

    from datetime import timedelta
    dates = []
    d = date_from
    while d <= date_to:
        if not weekdays_only or d.weekday() < 5:
            dates.append(d)
        d += timedelta(days=1)

    from services.roster_scheduler_service import bulk_fill
    try:
        result = bulk_fill(req_team_id, account_id, shift_code, dates, member_ids,
                           created_by_id=current_user.id)
        return jsonify({'success': True, **result})
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# API — reset month (DELETE non-protected rows)
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/schedule', methods=['DELETE'])
@login_required
def api_reset_schedule():
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()
    body = request.get_json(force=True) or {}

    try:
        req_team_id = int(body.get('team_id', team_id))
        year = _clamp_year(int(body.get('year', date.today().year)))
        month = int(body.get('month', date.today().month))
    except (ValueError, TypeError) as exc:
        return jsonify({'success': False, 'error': f'Invalid parameters: {exc}'}), 400

    from services.roster_scheduler_service import reset_month_schedule
    try:
        result = reset_month_schedule(req_team_id, account_id, year, month)
        return jsonify({'success': True, **result})
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# API — sync to existing ShiftRoster table
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/sync-to-roster', methods=['POST'])
@login_required
def api_sync_to_roster():
    """
    Push the auto-generated schedule for a month into the main ShiftRoster table
    so it appears in the existing Shift Roster tab.
    """
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()
    body = request.get_json(force=True) or {}

    try:
        req_team_id = int(body.get('team_id', team_id))
        year        = int(body.get('year',    date.today().year))
        month       = int(body.get('month',   date.today().month))
        overwrite   = bool(body.get('overwrite', False))
    except (ValueError, TypeError) as exc:
        return jsonify({'success': False, 'error': f'Invalid parameters: {exc}'}), 400

    from services.roster_scheduler_service import sync_to_shift_roster
    from services.email_service import send_schedule_published_notification
    try:
        result = sync_to_shift_roster(req_team_id, account_id, year, month, overwrite)
        # Fire-and-forget notification — failure should not block the response
        try:
            send_schedule_published_notification(
                team_id=req_team_id,
                year=year,
                month=month,
                published_by=current_user.username or current_user.name or 'Admin',
                synced_count=result.get('synced', 0),
            )
        except Exception:
            pass
        return jsonify({'success': True, **result})
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# API — export to XLSX
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/export')
@login_required
def api_export_schedule():
    account_id, team_id = _current_account_team()
    year, month = _parse_year_month(request)

    req_team_id = request.args.get('team_id', team_id)
    try:
        req_team_id = int(req_team_id)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid team_id'}), 400

    from services.roster_scheduler_service import get_shift_view, export_to_xlsx
    try:
        data = get_shift_view(req_team_id, account_id, year, month)
        xlsx_bytes = export_to_xlsx(data, year, month)
        import calendar
        month_name = calendar.month_name[month]
        filename = f'roster_{year}_{month:02d}.xlsx'
        from flask import send_file
        import io
        return send_file(
            io.BytesIO(xlsx_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# API — coverage view
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/coverage')
@login_required
def api_get_coverage():
    account_id, team_id = _current_account_team()
    year, month = _parse_year_month(request)

    req_team_id = request.args.get('team_id', team_id)
    try:
        req_team_id = int(req_team_id)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid team_id'}), 400

    from services.roster_scheduler_service import get_coverage_view
    try:
        data = get_coverage_view(req_team_id, account_id, year, month)
        return jsonify({'success': True, **data})
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# API — member summary
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/summary')
@login_required
def api_get_summary():
    account_id, team_id = _current_account_team()
    year, month = _parse_year_month(request)

    req_team_id = request.args.get('team_id', team_id)
    try:
        req_team_id = int(req_team_id)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid team_id'}), 400

    from services.roster_scheduler_service import get_member_summary
    try:
        data = get_member_summary(req_team_id, account_id, year, month)
        return jsonify({'success': True, 'data': data})
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# API — CSV import
# ---------------------------------------------------------------------------

@roster_scheduler_bp.route('/api/roster/import', methods=['POST'])
@login_required
def api_import_csv():
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith('.csv'):
        return jsonify({'success': False, 'error': 'Please upload a .csv file'}), 400
    f.stream.seek(0, 2)
    if f.stream.tell() > 5 * 1024 * 1024:
        return jsonify({'success': False, 'error': 'File too large. Maximum size is 5 MB.'}), 413
    f.stream.seek(0)

    try:
        req_team_id = int(request.form.get('team_id', team_id))
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid team_id'}), 400

    from services.roster_scheduler_service import import_from_csv
    try:
        result = import_from_csv(f.stream, req_team_id, account_id, created_by_id=current_user.id)
        return jsonify({'success': True, **result})
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


@roster_scheduler_bp.route('/api/roster/upload-leave-plan', methods=['POST'])
@login_required
def api_upload_leave_plan():
    if not _admin_required():
        return jsonify({'success': False, 'error': 'Admin required'}), 403

    account_id, team_id = _current_account_team()

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Please upload an .xlsx or .xls file'}), 400
    f.stream.seek(0, 2)
    if f.stream.tell() > 5 * 1024 * 1024:
        return jsonify({'success': False, 'error': 'File too large. Maximum size is 5 MB.'}), 413
    f.stream.seek(0)

    try:
        req_team_id = int(request.form.get('team_id', team_id))
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid team_id'}), 400

    leave_type = request.form.get('leave_type', 'VL').upper()

    from services.roster_scheduler_service import upload_leave_plan
    try:
        result = upload_leave_plan(f.stream, req_team_id, account_id, leave_type,
                                    created_by_id=current_user.id)
        return jsonify({'success': True, **result})
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500


@roster_scheduler_bp.route('/api/roster/leave-plan-sample', methods=['GET'])
@login_required
def api_leave_plan_sample():
    """Generate and return a sample leave plan Excel file."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from io import BytesIO
    from flask import send_file
    from datetime import date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leave Plan'

    today = date.today()
    months = []
    for i in range(6):
        m = today.month + i
        y = today.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        abbr = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'][m - 1]
        months.append(f'{abbr}-{str(y)[2:]}')

    headers = ['Resource Name', 'Stream'] + months
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    sample_rows = [
        ['John Smith',   'NOC L1', '3,4,5',  '',       '12,13', '',      '',      ''],
        ['Jane Doe',     'NOC L2', '',        '1-5',    '',      '20,21', '',      ''],
        ['Alex Johnson', 'NOC L1', '10,11',   '',       '',      '',      '1-3',   ''],
    ]

    alt_fill = PatternFill('solid', fgColor='EBF3FB')
    for row_idx, row_data in enumerate(sample_rows, 2):
        row_data_padded = row_data + [''] * (len(headers) - len(row_data))
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row_data_padded[:len(headers)], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left', vertical='center')
            cell.border = border
            if fill:
                cell.fill = fill

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
    ws.row_dimensions[1].height = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='leave_plan_sample.xlsx'
    )
