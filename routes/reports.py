
from flask import Blueprint, render_template, request, send_file, session, jsonify, flash, redirect, url_for
from flask_login import login_required, current_user
from datetime import datetime
from sqlalchemy import func
from collections import OrderedDict
from models.models import Shift, Incident, ShiftKeyPoint, TeamMember, Account, Team, User, ShiftChangeInfo, ShiftKBUpdate, db
from models.audit_log import AuditLog
from services.export_service import export_incidents_csv, export_keypoints_pdf
from services.team_access_service import TeamAccessService
from services.multi_team_service import MultiTeamService
from services.audit_service import log_action
import logging

# Module logger
logger = logging.getLogger(__name__)

def format_excel_sheet(worksheet, df):
    """Apply professional formatting to Excel worksheet"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Define styles
    header_fill = PatternFill(start_color='6B8DD6', end_color='6B8DD6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # Status colors
    status_colors = {
        'Open': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
        'In Progress': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
        'Closed': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
        'Completed': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
        'Published': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
        'New': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
        'In Review': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
        'Cancelled': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
        'Scheduled': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
        'Postponed': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
    }
    
    # Format header row
    for col_idx, col in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Format data rows
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Apply status colors
            cell_value = str(cell.value) if cell.value else ''
            if cell_value in status_colors:
                cell.fill = status_colors[cell_value]
        
        # Alternate row coloring
        if row_idx % 2 == 0:
            alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if not cell.fill or cell.fill.fgColor.rgb == '00000000':
                    cell.fill = alt_fill
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(df.columns, 1):
        max_length = max(
            df[col].astype(str).map(len).max() if len(df) > 0 else 0,
            len(str(col))
        )
        adjusted_width = min(max_length + 3, 50)
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = adjusted_width
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 25
    
    # Freeze the header row
    worksheet.freeze_panes = 'A2'

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/reports', methods=['GET'])
@login_required
def reports():
    """Main reports page - redirects to handover reports"""
    log_action('View Reports Tab', 'Accessed main reports page')
    # Redirect to handover reports as the main reports page
    from flask import redirect, url_for
    return redirect(url_for('reports.handover_reports'))


@reports_bp.route('/change-info-reports', methods=['GET'])
@login_required
def change_info_reports():
    """Change Info reports page"""
    log_action('View Change Info Reports', 'Accessed change info reports page')
    
    # Get filter parameters
    date_filter = request.args.get('date')
    account_id_filter = request.args.get('account_id')
    team_id_filter = request.args.get('team_id')
    app_filter = request.args.get('app_name')
    status_filter = request.args.get('status')
    
    # Initialize team context variables
    selected_team_id = None
    all_teams_selected = False
    team_filter_context = None
    accounts = []
    teams = []
    
    # Build query based on user permissions - select both tables
    query = db.session.query(ShiftChangeInfo, Shift.date.label('shift_date')).join(Shift)
    
    if current_user.role == 'super_admin':
        # Super admin can see all change info
        accounts = Account.query.filter_by(status='active').all()
        teams = Team.query.filter_by(status='active')
        if account_id_filter:
            teams = teams.filter_by(account_id=account_id_filter)
        teams = teams.all()
        
        # Handle team filtering for super admin
        if team_id_filter == '':
            all_teams_selected = True
        elif team_id_filter:
            try:
                selected_team_id = int(team_id_filter)
            except (TypeError, ValueError):
                selected_team_id = None
                
    elif current_user.role == 'account_admin':
        # Account admin can see their account's change info
        query = query.filter(ShiftChangeInfo.account_id == current_user.account_id)
        accounts = [Account.query.get(current_user.account_id)] if current_user.account_id else []
        teams = Team.query.filter_by(account_id=current_user.account_id, status='active').all()
        
        # Handle team filtering for account admin
        if team_id_filter == '':
            all_teams_selected = True
        elif team_id_filter:
            try:
                selected_team_id = int(team_id_filter)
            except (TypeError, ValueError):
                selected_team_id = None
                
    else:
        # Regular users: use team access service filtering
        team_filter_context = TeamAccessService.get_team_filter_context()
        teams = team_filter_context['user_teams']
        accounts = [Account.query.get(current_user.account_id)] if current_user.account_id else []
        
        # Handle team selection for regular users
        team_param = request.args.get('team_id')
        all_teams_selected = team_param == ''
        
        if team_param is None:
            # No filter provided -> default to previously selected/primary team
            selected_team_id = team_filter_context.get('selected_team_id')
        elif team_param == '':
            # Explicit "All Teams" selection
            selected_team_id = None
            all_teams_selected = True
        else:
            # Specific team selected
            try:
                selected_team_id = int(team_param)
            except (TypeError, ValueError):
                selected_team_id = team_filter_context.get('selected_team_id')
        
        # Apply team filtering for regular users
        if all_teams_selected:
            # Show all teams user has access to
            query = TeamAccessService.apply_team_filter(query, ShiftChangeInfo)
        elif selected_team_id:
            # Filter by specific team
            query = query.filter(ShiftChangeInfo.team_id == selected_team_id)
        else:
            # Fallback to team access service filtering
            query = TeamAccessService.apply_team_filter(query, ShiftChangeInfo)
    
    # Apply filters
    if date_filter:
        try:
            date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
            query = query.filter(Shift.date == date_obj)
        except ValueError:
            pass
    
    if account_id_filter and current_user.role == 'super_admin':
        query = query.filter(ShiftChangeInfo.account_id == account_id_filter)
    
    if team_id_filter and current_user.role in ['super_admin', 'account_admin'] and not all_teams_selected:
        if selected_team_id:
            query = query.filter(ShiftChangeInfo.team_id == selected_team_id)
    
    if app_filter:
        query = query.filter(ShiftChangeInfo.app_name.ilike(f'%{app_filter}%'))
    
    if status_filter:
        query = query.filter(ShiftChangeInfo.status == status_filter)
    
    # Order by most recent first (MySQL doesn't support NULLS LAST)
    raw_change_infos = query.order_by(Shift.date.desc(), ShiftChangeInfo.change_datetime.desc()).all()
    
    # Deduplicate by change_number OR app_name+description - keep most recent version of each unique change
    change_map = {}
    for change_info, shift_date in raw_change_infos:
        # Use change_number if available, otherwise use app_name + description hash
        if change_info.change_number and change_info.change_number.strip():
            change_key = change_info.change_number.strip().lower()
        else:
            # Use a combination of app_name and description for deduplication if change_number is missing
            change_key = f"{change_info.app_name or ''}_{(change_info.description or '')[:50]}".strip().lower()
        
        if change_key and (change_key not in change_map or change_info.id > change_map[change_key][0].id):
            change_map[change_key] = (change_info, shift_date)
    
    change_infos = list(change_map.values())
    
    return render_template('change_info_reports.html',
                         change_infos=change_infos,
                         accounts=accounts,
                         teams=teams,
                         date_filter=date_filter,
                         account_id_filter=account_id_filter,
                         team_id_filter=team_id_filter,
                         selected_team_id=selected_team_id,
                         all_teams_selected=all_teams_selected,
                         app_filter=app_filter,
                         status_filter=status_filter)


@reports_bp.route('/kb-update-reports', methods=['GET'])
@login_required
def kb_update_reports():
    """KB Update reports page"""
    log_action('View KB Update Reports', 'Accessed KB update reports page')
    
    # Get filter parameters
    date_filter = request.args.get('date')
    account_id_filter = request.args.get('account_id')
    team_id_filter = request.args.get('team_id')
    app_filter = request.args.get('app_name')
    status_filter = request.args.get('status')
    
    # Initialize team context variables
    selected_team_id = None
    all_teams_selected = False
    team_filter_context = None
    accounts = []
    teams = []
    
    # Build query based on user permissions - select both tables
    query = db.session.query(ShiftKBUpdate, Shift.date.label('shift_date')).join(Shift)
    
    if current_user.role == 'super_admin':
        # Super admin can see all KB updates
        accounts = Account.query.filter_by(status='active').all()
        teams = Team.query.filter_by(status='active')
        if account_id_filter:
            teams = teams.filter_by(account_id=account_id_filter)
        teams = teams.all()
        
        # Handle team filtering for super admin
        if team_id_filter == '':
            all_teams_selected = True
        elif team_id_filter:
            try:
                selected_team_id = int(team_id_filter)
            except (TypeError, ValueError):
                selected_team_id = None
                
    elif current_user.role == 'account_admin':
        # Account admin can see their account's KB updates
        query = query.filter(ShiftKBUpdate.account_id == current_user.account_id)
        accounts = [Account.query.get(current_user.account_id)] if current_user.account_id else []
        teams = Team.query.filter_by(account_id=current_user.account_id, status='active').all()
        
        # Handle team filtering for account admin
        if team_id_filter == '':
            all_teams_selected = True
        elif team_id_filter:
            try:
                selected_team_id = int(team_id_filter)
            except (TypeError, ValueError):
                selected_team_id = None
                
    else:
        # Regular users: use team access service filtering
        team_filter_context = TeamAccessService.get_team_filter_context()
        teams = team_filter_context['user_teams']
        accounts = [Account.query.get(current_user.account_id)] if current_user.account_id else []
        
        # Handle team selection for regular users
        team_param = request.args.get('team_id')
        all_teams_selected = team_param == ''
        
        if team_param is None:
            # No filter provided -> default to previously selected/primary team
            selected_team_id = team_filter_context.get('selected_team_id')
        elif team_param == '':
            # Explicit "All Teams" selection
            selected_team_id = None
            all_teams_selected = True
        else:
            # Specific team selected
            try:
                selected_team_id = int(team_param)
            except (TypeError, ValueError):
                selected_team_id = team_filter_context.get('selected_team_id')
        
        # Apply team filtering for regular users
        if all_teams_selected:
            # Show all teams user has access to
            query = TeamAccessService.apply_team_filter(query, ShiftKBUpdate)
        elif selected_team_id:
            # Filter by specific team
            query = query.filter(ShiftKBUpdate.team_id == selected_team_id)
        else:
            # Fallback to team access service filtering
            query = TeamAccessService.apply_team_filter(query, ShiftKBUpdate)
    
    # Apply filters
    if date_filter:
        try:
            date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
            query = query.filter(Shift.date == date_obj)
        except ValueError:
            pass
    
    if account_id_filter and current_user.role == 'super_admin':
        query = query.filter(ShiftKBUpdate.account_id == account_id_filter)
    
    if team_id_filter and current_user.role in ['super_admin', 'account_admin'] and not all_teams_selected:
        if selected_team_id:
            query = query.filter(ShiftKBUpdate.team_id == selected_team_id)
    
    if app_filter:
        query = query.filter(ShiftKBUpdate.app_name.ilike(f'%{app_filter}%'))
    
    if status_filter:
        query = query.filter(ShiftKBUpdate.status == status_filter)
    
    # Order by most recent first
    raw_kb_updates = query.order_by(Shift.date.desc(), ShiftKBUpdate.id.desc()).all()
    
    # 🔧 FIX: Improved deduplication - keep most recent version of each unique KB
    # Handle KBs with and without kb_number
    kb_map = {}
    for kb_update, shift_date in raw_kb_updates:
        # Use kb_number if available, otherwise use app_name + description hash (like carryforward)
        if kb_update.kb_number and kb_update.kb_number.strip():
            kb_key = kb_update.kb_number.strip().lower()
        else:
            kb_key = f"{kb_update.app_name or ''}_{(kb_update.description or '')[:50]}".strip().lower()
        
        # Keep the most recent version (highest ID) of each unique KB
        if kb_key and (kb_key not in kb_map or kb_update.id > kb_map[kb_key][0].id):
            kb_map[kb_key] = (kb_update, shift_date)
    
    kb_updates = list(kb_map.values())
    
    # Get unique statuses for filter dropdown
    status_options = ['New', 'Draft', 'In Review', 'Published']
    
    return render_template('kb_update_reports.html',
                         kb_updates=kb_updates,
                         accounts=accounts,
                         teams=teams,
                         date_filter=date_filter,
                         account_id_filter=account_id_filter,
                         team_id_filter=team_id_filter,
                         selected_team_id=selected_team_id,
                         all_teams_selected=all_teams_selected,
                         app_filter=app_filter,
                         status_filter=status_filter,
                         status_options=status_options)


# Bulk export filtered handover reports as CSV or PDF
@reports_bp.route('/handover-reports/export/bulk', methods=['GET'])
@login_required
def export_handover_bulk():
    log_action('Export Reports', f'Format: {request.args.get("format")}, Filters: account_id={request.args.get("account_id")}, team_id={request.args.get("team_id")}, date={request.args.get("date")}, shift_type={request.args.get("shift_type")}')
    date_filter = request.args.get('date')
    shift_type_filter = request.args.get('shift_type')
    account_id = request.args.get('account_id')
    team_id = request.args.get('team_id')
    format_type = request.args.get('format', 'csv')
    query = Shift.query
    
    # Apply team access filtering based on user role
    if current_user.role == 'super_admin':
        # Super admin can see all data - apply explicit filters only
        if account_id:
            query = query.filter_by(account_id=account_id)
        if team_id:
            query = query.filter_by(team_id=team_id)
    elif current_user.role == 'account_admin':
        # Account admin sees all teams in their account
        query = query.filter_by(account_id=current_user.account_id)
        if team_id:
            query = query.filter_by(team_id=team_id)
    else:
        # Regular users: restrict to their accessible teams only
        effective_team_ids = TeamAccessService.get_effective_team_ids()
        if effective_team_ids:
            if team_id:
                # If a specific team is requested, verify user has access
                try:
                    requested_team_id = int(team_id)
                    if requested_team_id in effective_team_ids:
                        query = query.filter_by(team_id=requested_team_id)
                    else:
                        # User doesn't have access to this team, filter by their teams
                        query = query.filter(Shift.team_id.in_(effective_team_ids))
                except (TypeError, ValueError):
                    query = query.filter(Shift.team_id.in_(effective_team_ids))
            else:
                # No specific team requested, show all user's accessible teams
                query = query.filter(Shift.team_id.in_(effective_team_ids))
        else:
            # No teams accessible, return empty
            query = query.filter(Shift.team_id == -1)  # This will return no results
        
        # Also filter by account
        if current_user.account_id:
            query = query.filter_by(account_id=current_user.account_id)
    
    # Apply date and shift type filters
    if date_filter:
        try:
            date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
            query = query.filter_by(date=date_obj)
        except Exception:
            pass
    if shift_type_filter:
        query = query.filter_by(current_shift_type=shift_type_filter)
    
    shifts = query.order_by(Shift.date.desc()).all()
    logger.info(f"Export: User {current_user.username} (role: {current_user.role}) exporting {len(shifts)} shifts")
    
    # Generate filename with date
    from datetime import datetime as dt
    export_date = dt.now().strftime('%Y%m%d_%H%M')
    
    # Collect detailed data for each category
    summary_rows = []
    incident_rows = []
    keypoint_rows = []
    change_rows = []
    kb_rows = []
    
    for shift in shifts:
        incidents = Incident.query.filter_by(shift_id=shift.id).all()
        key_points = ShiftKeyPoint.query.filter_by(shift_id=shift.id).all()
        # 🔧 FIX: Query change_infos and kb_updates by date/team/account with deduplication
        from datetime import timedelta
        raw_change_infos = ShiftChangeInfo.query.filter(
            ShiftChangeInfo.account_id == shift.account_id,
            ShiftChangeInfo.team_id == shift.team_id,
            ShiftChangeInfo.created_at >= shift.date,
            ShiftChangeInfo.created_at < shift.date + timedelta(days=1),
            ~ShiftChangeInfo.status.in_(['Completed', 'Cancelled', 'Implemented'])
        ).order_by(ShiftChangeInfo.id.desc()).all()
        # Deduplicate by change_number
        change_map = {}
        for ci in raw_change_infos:
            key = ci.change_number.strip().lower() if ci.change_number else f"{ci.app_name}_{ci.description[:30] if ci.description else ''}"
            if key not in change_map:
                change_map[key] = ci
        change_infos = list(change_map.values())
        
        raw_kb_updates = ShiftKBUpdate.query.filter(
            ShiftKBUpdate.account_id == shift.account_id,
            ShiftKBUpdate.team_id == shift.team_id,
            ShiftKBUpdate.created_at >= shift.date,
            ShiftKBUpdate.created_at < shift.date + timedelta(days=1),
            ShiftKBUpdate.status != 'Published'
        ).order_by(ShiftKBUpdate.id.desc()).all()
        # Deduplicate by kb_number
        kb_map = {}
        for kb in raw_kb_updates:
            key = kb.kb_number.strip().lower() if kb.kb_number else f"{kb.app_name}_{kb.description[:30] if kb.description else ''}"
            if key not in kb_map:
                kb_map[key] = kb
        kb_updates = list(kb_map.values())
        
        # Get team name
        team_name = ''
        if shift.team_id:
            team = Team.query.get(shift.team_id)
            if team:
                team_name = team.name
        
        # Find who submitted this handover
        submitted_by = 'Unknown'
        audit_entry = AuditLog.query.filter(
            AuditLog.action.like('%Create Handover%'),
            AuditLog.details.like(f'%Shift: {shift.current_shift_type}%'),
            AuditLog.details.like(f'%Date: {shift.date}%')
        ).first()
        
        if audit_entry:
            if audit_entry.user_id:
                user = User.query.get(audit_entry.user_id)
                if user:
                    submitted_by = user.display_name
                else:
                    submitted_by = audit_entry.username or 'Unknown User'
            else:
                submitted_by = audit_entry.username or 'Unknown User'
        
        # Summary row
        summary_rows.append({
            'Date': shift.date.strftime('%Y-%m-%d') if shift.date else '',
            'Team': team_name,
            'Shift': f"{shift.current_shift_type} → {shift.next_shift_type}",
            'Status': shift.status or 'Completed',
            'Submitted By': submitted_by,
            'Total Incidents': len(incidents),
            'Total Key Points': len(key_points),
            'Total Changes': len(change_infos),
            'Total KB Updates': len(kb_updates)
        })
        
        # Detailed incident rows
        for idx, inc in enumerate(incidents, 1):
            incident_rows.append({
                'Date': shift.date.strftime('%Y-%m-%d') if shift.date else '',
                'Shift': f"{shift.current_shift_type} → {shift.next_shift_type}",
                'Team': team_name,
                '#': idx,
                'Type': inc.type or 'N/A',
                'Incident ID': inc.title.split(' - ')[1] if ' - ' in (inc.title or '') else (inc.title or 'N/A'),
                'Application': inc.title.split(' - ')[0] if ' - ' in (inc.title or '') else 'N/A',
                'Status': inc.status or 'N/A',
                'Priority': inc.priority or 'N/A',
                'Escalated To': inc.escalated_to or '-',
                'Handover Notes': inc.handover or ''
            })
        
        # Detailed key point rows
        for idx, kp in enumerate(key_points, 1):
            responsible_name = 'N/A'
            if kp.responsible_engineer_id:
                engineer = TeamMember.query.get(kp.responsible_engineer_id)
                if engineer:
                    responsible_name = engineer.name
            
            # Get latest update for this key point
            latest_update = ''
            if kp.updates:
                sorted_updates = sorted(kp.updates, key=lambda u: u.update_date if u.update_date else '', reverse=True)
                if sorted_updates:
                    latest_update = sorted_updates[0].update_text or ''
            
            keypoint_rows.append({
                'Date': shift.date.strftime('%Y-%m-%d') if shift.date else '',
                'Shift': f"{shift.current_shift_type} → {shift.next_shift_type}",
                'Team': team_name,
                '#': idx,
                'Description': kp.description or '',
                'Status': kp.status or 'N/A',
                'Responsible Engineer': responsible_name,
                'JIRA ID': kp.jira_id or '',
                'Latest Update': latest_update[:200] if latest_update else ''
            })
        
        # Detailed change info rows
        for idx, ci in enumerate(change_infos, 1):
            responsible_name = 'N/A'
            if ci.responsible_engineer_id:
                engineer = TeamMember.query.get(ci.responsible_engineer_id)
                if engineer:
                    responsible_name = engineer.name
            
            change_rows.append({
                'Date': shift.date.strftime('%Y-%m-%d') if shift.date else '',
                'Shift': f"{shift.current_shift_type} → {shift.next_shift_type}",
                'Team': team_name,
                '#': idx,
                'Change Number': ci.change_number or 'N/A',
                'Application': ci.app_name or '',
                'Description': ci.description or '',
                'Change Date': ci.change_datetime.strftime('%Y-%m-%d %H:%M') if ci.change_datetime else 'N/A',
                'Status': ci.status or 'N/A',
                'Responsible Engineer': responsible_name
            })
        
        # Detailed KB update rows
        for idx, kb in enumerate(kb_updates, 1):
            responsible_name = 'N/A'
            if kb.responsible_engineer_id:
                engineer = TeamMember.query.get(kb.responsible_engineer_id)
                if engineer:
                    responsible_name = engineer.name
            
            kb_rows.append({
                'Date': shift.date.strftime('%Y-%m-%d') if shift.date else '',
                'Shift': f"{shift.current_shift_type} → {shift.next_shift_type}",
                'Team': team_name,
                '#': idx,
                'KB Number': kb.kb_number or 'N/A',
                'Application': kb.app_name or '',
                'Description': kb.description or '',
                'Status': kb.status or 'N/A',
                'Responsible Engineer': responsible_name
            })
    
    if format_type == 'csv':
        # For CSV, create a detailed format grouped by DATE
        import pandas as pd, io
        from collections import defaultdict
        
        # Group all rows by date
        incidents_by_date = defaultdict(list)
        keypoints_by_date = defaultdict(list)
        changes_by_date = defaultdict(list)
        kb_by_date = defaultdict(list)
        
        for row in incident_rows:
            incidents_by_date[row['Date']].append(row)
        for row in keypoint_rows:
            keypoints_by_date[row['Date']].append(row)
        for row in change_rows:
            changes_by_date[row['Date']].append(row)
        for row in kb_rows:
            kb_by_date[row['Date']].append(row)
        
        # Get all unique dates and sort (newest first)
        all_dates = sorted(set(
            list(incidents_by_date.keys()) + 
            list(keypoints_by_date.keys()) + 
            list(changes_by_date.keys()) + 
            list(kb_by_date.keys())
        ), reverse=True)
        
        all_rows = []
        
        for date in all_dates:
            # Add date header
            all_rows.append({'Date': f'=== {date} ===' + '=' * 40, 'Type': '', 'Team': '', 'Shift': ''})
            
            # Add incidents for this date
            if date in incidents_by_date:
                for row in incidents_by_date[date]:
                    all_rows.append({
                        'Date': row['Date'],
                        'Type': 'INCIDENT',
                        'Team': row['Team'],
                        'Shift': row['Shift'],
                        '#': row['#'],
                        'Category': row['Type'],
                        'ID/Number': row['Incident ID'],
                        'Application': row['Application'],
                        'Description': '',
                        'Status': row['Status'],
                        'Priority': row['Priority'],
                        'Responsible/Escalated': row['Escalated To'],
                        'Notes': row['Handover Notes']
                    })
            
            # Add key points for this date
            if date in keypoints_by_date:
                for row in keypoints_by_date[date]:
                    all_rows.append({
                        'Date': row['Date'],
                        'Type': 'KEY POINT',
                        'Team': row['Team'],
                        'Shift': row['Shift'],
                        '#': row['#'],
                        'Category': '',
                        'ID/Number': row['JIRA ID'],
                        'Application': '',
                        'Description': row['Description'],
                        'Status': row['Status'],
                        'Priority': '',
                        'Responsible/Escalated': row['Responsible Engineer'],
                        'Notes': row['Latest Update']
                    })
            
            # Add changes for this date
            if date in changes_by_date:
                for row in changes_by_date[date]:
                    all_rows.append({
                        'Date': row['Date'],
                        'Type': 'CHANGE',
                        'Team': row['Team'],
                        'Shift': row['Shift'],
                        '#': row['#'],
                        'Category': '',
                        'ID/Number': row['Change Number'],
                        'Application': row['Application'],
                        'Description': row['Description'],
                        'Status': row['Status'],
                        'Priority': '',
                        'Responsible/Escalated': row['Responsible Engineer'],
                        'Notes': row['Change Date']
                    })
            
            # Add KB updates for this date
            if date in kb_by_date:
                for row in kb_by_date[date]:
                    all_rows.append({
                        'Date': row['Date'],
                        'Type': 'KB UPDATE',
                        'Team': row['Team'],
                        'Shift': row['Shift'],
                        '#': row['#'],
                        'Category': '',
                        'ID/Number': row['KB Number'],
                        'Application': row['Application'],
                        'Description': row['Description'],
                        'Status': row['Status'],
                        'Priority': '',
                        'Responsible/Escalated': row['Responsible Engineer'],
                        'Notes': ''
                    })
            
            # Add blank row after each date
            all_rows.append({})
        
        df = pd.DataFrame(all_rows)
        csv_io = io.StringIO()
        df.to_csv(csv_io, index=False)
        csv_io.seek(0)
        return send_file(
            io.BytesIO(csv_io.getvalue().encode('utf-8-sig')),
            mimetype='text/csv', 
            as_attachment=True, 
            download_name=f'handover_reports_detailed_{export_date}.csv'
        )
    
    elif format_type == 'excel':
        import pandas as pd, io
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        excel_io = io.BytesIO()
        
        with pd.ExcelWriter(excel_io, engine='openpyxl') as writer:
            # Sheet 1: Summary
            if summary_rows:
                df_summary = pd.DataFrame(summary_rows)
                df_summary.to_excel(writer, index=False, sheet_name='Summary')
                format_excel_sheet(writer.sheets['Summary'], df_summary)
            
            # Sheet 2: Incidents (Detailed)
            if incident_rows:
                df_incidents = pd.DataFrame(incident_rows)
                df_incidents.to_excel(writer, index=False, sheet_name='Incidents')
                format_excel_sheet(writer.sheets['Incidents'], df_incidents)
            
            # Sheet 3: Key Points (Detailed)
            if keypoint_rows:
                df_keypoints = pd.DataFrame(keypoint_rows)
                df_keypoints.to_excel(writer, index=False, sheet_name='Key Points')
                format_excel_sheet(writer.sheets['Key Points'], df_keypoints)
            
            # Sheet 4: Change Info (Detailed)
            if change_rows:
                df_changes = pd.DataFrame(change_rows)
                df_changes.to_excel(writer, index=False, sheet_name='Change Info')
                format_excel_sheet(writer.sheets['Change Info'], df_changes)
            
            # Sheet 5: KB Updates (Detailed)
            if kb_rows:
                df_kb = pd.DataFrame(kb_rows)
                df_kb.to_excel(writer, index=False, sheet_name='KB Updates')
                format_excel_sheet(writer.sheets['KB Updates'], df_kb)
        
        excel_io.seek(0)
        return send_file(
            excel_io, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            as_attachment=True, 
            download_name=f'handover_reports_detailed_{export_date}.xlsx'
        )
    elif format_type == 'pdf':
        import io
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        
        pdf_io = io.BytesIO()
        doc = SimpleDocTemplate(pdf_io, pagesize=landscape(A4), 
                               rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=10, textColor=colors.HexColor('#6b8dd6'))
        section_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=14, spaceBefore=20, spaceAfter=10, textColor=colors.HexColor('#2c3e50'))
        cell_style = ParagraphStyle('cell', fontSize=8, leading=10)
        
        # Title
        elements.append(Paragraph("Shift Handover Reports - Detailed Export", title_style))
        elements.append(Paragraph(f"Generated on: {dt.now().strftime('%B %d, %Y at %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Helper function for creating styled tables
        def create_styled_table(data, col_widths, header_color='#6b8dd6'):
            if not data or len(data) <= 1:
                return None
            table = Table(data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(header_color)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))
            return table
        
        # 1. SUMMARY SECTION
        elements.append(Paragraph("1. SUMMARY", section_style))
        if summary_rows:
            summary_data = [['Date', 'Team', 'Shift', 'Status', 'Submitted By', 'Incidents', 'Key Points', 'Changes', 'KBs']]
            for row in summary_rows:
                summary_data.append([
                    row['Date'], row['Team'] or 'N/A', row['Shift'], row['Status'], 
                    row['Submitted By'], str(row['Total Incidents']), str(row['Total Key Points']),
                    str(row['Total Changes']), str(row['Total KB Updates'])
                ])
            table = create_styled_table(summary_data, [60, 80, 90, 50, 90, 50, 55, 50, 40])
            if table:
                elements.append(table)
        else:
            elements.append(Paragraph("No handover reports found.", styles['Normal']))
        
        # 2. INCIDENTS SECTION
        if incident_rows:
            elements.append(PageBreak())
            elements.append(Paragraph("2. INCIDENTS (Detailed)", section_style))
            inc_data = [['Date', 'Shift', '#', 'Type', 'Incident ID', 'Application', 'Status', 'Priority', 'Escalated To']]
            for row in incident_rows:
                status = row['Status']
                # Add visual indicator for status
                status_indicator = f"● {status}" if status in ['Open', 'In Progress'] else f"✓ {status}" if status == 'Closed' else status
                inc_data.append([
                    row['Date'], row['Shift'], str(row['#']), row['Type'],
                    Paragraph(str(row['Incident ID']), cell_style),
                    Paragraph(str(row['Application']), cell_style),
                    status_indicator, row['Priority'], row['Escalated To']
                ])
            table = create_styled_table(inc_data, [55, 80, 25, 50, 100, 100, 60, 50, 55], '#c62828')
            if table:
                elements.append(table)
        
        # 3. KEY POINTS SECTION
        if keypoint_rows:
            elements.append(PageBreak())
            elements.append(Paragraph("3. KEY POINTS (Detailed)", section_style))
            kp_data = [['Date', 'Shift', '#', 'Description', 'Status', 'Responsible', 'JIRA ID']]
            for row in keypoint_rows:
                status = row['Status']
                status_indicator = f"● {status}" if status in ['Open', 'In Progress'] else f"✓ {status}"
                kp_data.append([
                    row['Date'], row['Shift'], str(row['#']),
                    Paragraph(str(row['Description'])[:100] + ('...' if len(str(row['Description'])) > 100 else ''), cell_style),
                    status_indicator, row['Responsible Engineer'], row['JIRA ID'] or '-'
                ])
            table = create_styled_table(kp_data, [55, 80, 25, 250, 70, 90, 70], '#f57c00')
            if table:
                elements.append(table)
        
        # 4. CHANGE INFO SECTION
        if change_rows:
            elements.append(PageBreak())
            elements.append(Paragraph("4. CHANGE INFO (Detailed)", section_style))
            change_data = [['Date', 'Shift', '#', 'Change #', 'Application', 'Description', 'Status', 'Responsible']]
            for row in change_rows:
                change_data.append([
                    row['Date'], row['Shift'], str(row['#']), row['Change Number'],
                    Paragraph(str(row['Application']), cell_style),
                    Paragraph(str(row['Description'])[:80] + ('...' if len(str(row['Description'])) > 80 else ''), cell_style),
                    row['Status'], row['Responsible Engineer']
                ])
            table = create_styled_table(change_data, [55, 80, 25, 70, 90, 180, 60, 80], '#1565c0')
            if table:
                elements.append(table)
        
        # 5. KB UPDATES SECTION
        if kb_rows:
            elements.append(PageBreak())
            elements.append(Paragraph("5. KB UPDATES (Detailed)", section_style))
            kb_data = [['Date', 'Shift', '#', 'KB Number', 'Application', 'Description', 'Status', 'Responsible']]
            for row in kb_rows:
                kb_data.append([
                    row['Date'], row['Shift'], str(row['#']), row['KB Number'],
                    Paragraph(str(row['Application']), cell_style),
                    Paragraph(str(row['Description'])[:80] + ('...' if len(str(row['Description'])) > 80 else ''), cell_style),
                    row['Status'], row['Responsible Engineer']
                ])
            table = create_styled_table(kb_data, [55, 80, 25, 70, 90, 180, 60, 80], '#2e7d32')
            if table:
                elements.append(table)
        
        doc.build(elements)
        pdf_io.seek(0)
        return send_file(
            pdf_io, 
            mimetype='application/pdf', 
            as_attachment=True, 
            download_name=f'handover_reports_detailed_{export_date}.pdf'
        )
    else:
        return "Invalid format. Supported formats: csv, excel, pdf", 400


# Export incidents as CSV for a single shift
@reports_bp.route('/handover-reports/export/csv/<int:shift_id>', methods=['GET'])
@login_required
def export_handover_csv(shift_id):
    log_action('Export Single Shift CSV', f'Shift ID: {shift_id}')
    shift = Shift.query.get_or_404(shift_id)
    return export_incidents_csv(shift.date, shift_id)

# Export key points as PDF for a single shift
@reports_bp.route('/handover-reports/export/pdf/<int:shift_id>', methods=['GET'])
@login_required
def export_handover_pdf(shift_id):
    log_action('Export Single Shift PDF', f'Shift ID: {shift_id}')
    shift = Shift.query.get_or_404(shift_id)
    return export_keypoints_pdf(shift.date, shift_id)


@reports_bp.route('/handover-reports/detailed/<int:shift_id>', methods=['GET'])
@login_required
def detailed_shift_report(shift_id):
    """Detailed view of a single shift report with enhanced UI sections"""
    log_action('View Detailed Shift Report', f'Shift ID: {shift_id}')
    
    try:
        # Get the shift
        shift = Shift.query.get_or_404(shift_id)
        
        # Check permissions
        if current_user.role == 'user':
            # Regular users can only see shifts from their teams
            user_team_ids = [team.id for team in current_user.get_teams()]
            if shift.team_id not in user_team_ids and shift.team_id != current_user.team_id:
                from flask import abort
                abort(403)
        elif current_user.role == 'account_admin':
            # Account admins can only see their account's shifts
            if shift.team.account_id != current_user.account_id:
                from flask import abort
                abort(403)
        # Super admins can see all shifts
        
        # Get all related data
        incidents = Incident.query.filter_by(shift_id=shift_id).all()
        
        # 🔍 ENHANCED KEY POINTS DEBUG: Check multiple possibilities
        key_points = ShiftKeyPoint.query.filter_by(shift_id=shift_id).all()
        
        # For draft shifts, also check active key points from reports logic
        if shift.status == 'draft':
            logger.debug(f"🔍 DRAFT SHIFT DETECTED - checking active key points as well")
            # Get active key points like in main reports
            all_active_kps = ShiftKeyPoint.query.filter(
                ShiftKeyPoint.account_id == shift.account_id,
                ShiftKeyPoint.team_id == shift.team_id,
                ShiftKeyPoint.status.in_(['Open', 'In Progress'])
            ).all()
            logger.debug(f"🔍 Found {len(all_active_kps)} active key points for account/team")
            
            # Apply COMPLETE dashboard-style filtering and deduplication
            all_key_points = key_points + all_active_kps
            
            # First: Apply submission filtering like dashboard
            filtered_kps = []
            for kp in all_key_points:
                shift = Shift.query.get(kp.shift_id)
                if shift and shift.status == 'sent':
                    # Check if there's a newer version that closed this key point (CASE-INSENSITIVE)
                    newer_closed = ShiftKeyPoint.query.filter(
                        func.lower(ShiftKeyPoint.description) == kp.description.lower(),
                        ShiftKeyPoint.jira_id == kp.jira_id,
                        ShiftKeyPoint.status == 'Closed',
                        ShiftKeyPoint.id > kp.id
                    ).first()
                    if newer_closed:
                        logger.debug(f"🔍 DETAILED: Excluding key point ID {kp.id} - found newer closed version ID {newer_closed.id}")
                        continue
                filtered_kps.append(kp)
            
            # Second: Apply exact dashboard deduplication
            kp_map = {}
            for kp in filtered_kps:
                if kp.status == 'Closed':
                    continue
                normalized_jira = kp.jira_id if kp.jira_id and kp.jira_id.lower() not in ['none', 'null', ''] else None
                key = (kp.description, normalized_jira)
                if key not in kp_map or kp.id > kp_map[key].id:
                    kp_map[key] = kp
            
            key_points = list(kp_map.values())
            logger.debug(f"🔍 After dashboard-style filtering and deduplication: {len(key_points)} unique key points")
        
        # 🔧 FIX: Query change_infos by date/team/account AND deduplicate by change_number
        # This ensures ALL pending changes for this date appear, but without duplicates
        from datetime import timedelta
        raw_change_infos = ShiftChangeInfo.query.filter(
            ShiftChangeInfo.account_id == shift.account_id,
            ShiftChangeInfo.team_id == shift.team_id,
            ShiftChangeInfo.created_at >= shift.date,
            ShiftChangeInfo.created_at < shift.date + timedelta(days=1),
            ~ShiftChangeInfo.status.in_(['Completed', 'Cancelled', 'Implemented'])
        ).order_by(ShiftChangeInfo.id.desc()).all()
        
        # Deduplicate by change_number - keep most recent version
        change_map = {}
        for ci in raw_change_infos:
            key = ci.change_number.strip().lower() if ci.change_number else f"{ci.app_name}_{ci.description[:30]}"
            if key not in change_map:
                change_map[key] = ci
        change_infos = list(change_map.values())
        
        raw_kb_updates = ShiftKBUpdate.query.filter(
            ShiftKBUpdate.account_id == shift.account_id,
            ShiftKBUpdate.team_id == shift.team_id,
            ShiftKBUpdate.created_at >= shift.date,
            ShiftKBUpdate.created_at < shift.date + timedelta(days=1),
            ShiftKBUpdate.status != 'Published'
        ).order_by(ShiftKBUpdate.id.desc()).all()
        
        # Deduplicate by kb_number
        kb_map = {}
        for kb in raw_kb_updates:
            key = kb.kb_number.strip().lower() if kb.kb_number else f"{kb.app_name}_{kb.description[:30]}"
            if key not in kb_map:
                kb_map[key] = kb
        kb_updates = list(kb_map.values())
        
        # 🔍 DEBUG: Check what data we're getting for the detailed report
        logger.debug(f"🔍 DETAILED REPORT DEBUG for shift {shift_id}:")
        logger.debug(f"  - Shift status: {shift.status}")
        logger.debug(f"  - Incidents found: {len(incidents)}")
        logger.debug(f"  - Key Points found (after processing): {len(key_points)}")
        logger.debug(f"  - Change Infos found (excluding Completed/Cancelled): {len(change_infos)}")
        logger.debug(f"  - KB Updates found (excluding Published): {len(kb_updates)}")
        
        # Debug first few key points
        for i, kp in enumerate(key_points[:3]):
            logger.debug(f"    KP {i+1}: '{kp.description[:50]}...' (Status: {kp.status}, Shift: {kp.shift_id})")
        
        for i, change in enumerate(change_infos):
            logger.debug(f"    Change {i+1}: {change.app_name} - {change.status}")
        
        for i, kb in enumerate(kb_updates):
            logger.debug(f"    KB Update {i+1}: {kb.app_name} - {kb.status}")
        
        # Get team member info - use same logic as main reports route
        from models.models import User
        from models.handover_enhanced import HandoverRequest
        
        submitted_by = 'Unknown'
        try:
            # First try to find from HandoverRequest table using shift attributes
            handover_req = HandoverRequest.query.filter_by(
                shift_date=shift.date,
                current_shift_type=shift.current_shift_type,
                account_id=shift.account_id,
                team_id=shift.team_id
            ).first()
            
            if handover_req and handover_req.created_by_id:
                user = User.query.get(handover_req.created_by_id)
                if user:
                    submitted_by = user.display_name or user.username
                    logger.debug(f"🔍 DETAILED: Found accurate submitter: {submitted_by} (ID: {user.id}) for shift {shift.id}")
                else:
                    submitted_by = f'User ID: {handover_req.created_by_id}'
                    logger.debug(f"🔍 DETAILED: User not found, using ID: {submitted_by} for shift {shift.id}")
            else:
                # Fallback to audit logs
                logger.debug(f"🔍 DETAILED: No handover request found, checking audit logs for shift {shift.id}")
                audit_entry = AuditLog.query.filter(
                    AuditLog.action.contains('Handover Submitted'),
                    AuditLog.details.contains(f'Shift ID: {shift.id}')
                ).order_by(AuditLog.timestamp.desc()).first()
                
                if audit_entry:
                    if audit_entry.user_id:
                        user = User.query.get(audit_entry.user_id)
                        if user:
                            submitted_by = user.display_name or user.username
                            logger.debug(f"🔍 DETAILED: Fallback audit submitter: {submitted_by} for shift {shift.id}")
                        else:
                            submitted_by = audit_entry.username or 'Unknown User'
                            logger.debug(f"🔍 DETAILED: Using audit username: {submitted_by} for shift {shift.id}")
                else:
                    logger.debug(f"🔍 DETAILED: No audit entry found for shift {shift.id}")
                    submitted_by = 'Unknown'
        except Exception as e:
            logger.debug(f"🔍 DETAILED: Error finding submitter for shift {shift.id}: {e}")
            submitted_by = 'Unknown'
        
        # Organize incidents by type
        incidents_by_type = {
            'open': [],
            'closed': [],
            'priority': [],
            'handover': [],
            'escalated': []
        }
        
        for inc in incidents:
            incident_type = inc.type.lower() if inc.type else 'other'
            if incident_type in incidents_by_type:
                incidents_by_type[incident_type].append(inc)
            else:
                incidents_by_type.setdefault('other', []).append(inc)
        
        # Organize key points by status
        key_points_by_status = {
            'open': [],
            'in_progress': [],
            'closed': []
        }
        
        for kp in key_points:
            status = kp.status.lower().replace(' ', '_') if kp.status else 'open'
            if status in key_points_by_status:
                key_points_by_status[status].append(kp)
            else:
                key_points_by_status['open'].append(kp)
        
        return render_template(
            'detailed_shift_report.html',
            shift=shift,
            submitted_by=submitted_by,
            incidents=incidents,
            incidents_by_type=incidents_by_type,
            key_points=key_points,
            key_points_by_status=key_points_by_status,
            change_infos=change_infos,
            kb_updates=kb_updates
        )
        
    except Exception as e:
        logger.error(f"❌ Error in detailed_shift_report: {e}")
        from flask import abort
        abort(500)


@reports_bp.route('/handover-reports', methods=['GET'])
@login_required
def handover_reports():
    logger.debug(f"🚨 HANDOVER REPORTS ROUTE CALLED 🚨")
    
    # Check database integrity first
    total_shifts = Shift.query.count()
    total_incidents = Incident.query.count()
    total_key_points = ShiftKeyPoint.query.count()
    logger.debug(f"🔍 Database totals: {total_shifts} shifts, {total_incidents} incidents, {total_key_points} key points")
    
    # Show some sample records
    if total_incidents > 0:
        sample_incidents = Incident.query.limit(3).all()
        for inc in sample_incidents:
            logger.debug(f"🔍 Sample Incident: ID={inc.id}, shift_id={inc.shift_id}, title='{inc.title}'")
    
    if total_key_points > 0:
        sample_kps = ShiftKeyPoint.query.limit(3).all()
        for kp in sample_kps:
            logger.debug(f"🔍 Sample KeyPoint: ID={kp.id}, shift_id={kp.shift_id}, desc='{kp.description[:50]}'")
    
    try:
        # Default filters
        log_action('View Reports Tab', f'Filters: account_id={request.args.get("account_id")}, team_id={request.args.get("team_id")}, date={request.args.get("date")}, shift_type={request.args.get("shift_type")}')
        date_filter = request.args.get('date')
        shift_type_filter = request.args.get('shift_type')
        account_id = None
        team_id = None
        selected_team_id = None
        all_teams_selected = False
        team_filter_context = None
        accounts = []
        teams = []
        query = Shift.query
        if current_user.role == 'super_admin':
            accounts = Account.query.filter_by(is_active=True).all()
            # 🔧 FIX: For super_admin, only use filters if explicitly provided in request params
            # Don't use session values as defaults - this was causing empty results
            account_id = request.args.get('account_id')  # Don't use session defaults for super_admin
            logger.debug(f"🔍 SUPER_ADMIN: account_id from params only: {account_id}")
            
            # 🔧 FIX: Clear session defaults for super_admin to prevent stale selections
            if not account_id and 'selected_account_id' in session:
                session.pop('selected_account_id', None)
                logger.debug(f"🔍 SUPER_ADMIN: Cleared stale selected_account_id from session")
            teams = Team.query.filter_by(is_active=True)
            if account_id:
                teams = teams.filter_by(account_id=account_id)
                logger.debug(f"🔍 SUPER_ADMIN: Filtering teams by account_id: {account_id}")
            else:
                logger.debug(f"🔍 SUPER_ADMIN: No account_id in params - showing ALL accounts data")
            teams = teams.all()
            team_id = request.args.get('team_id')  # Don't use session defaults for super_admin
            if team_id == '':
                all_teams_selected = True
            elif team_id not in (None, 'None'):
                try:
                    selected_team_id = int(team_id)
                except (TypeError, ValueError):
                    selected_team_id = None
            logger.debug(f"🔍 SUPER_ADMIN: team_id from params only: {team_id}")
            
            # 🔧 FIX: Clear session defaults for super_admin to prevent stale selections  
            if not team_id and 'selected_team_id' in session:
                session.pop('selected_team_id', None)
                logger.debug(f"🔍 SUPER_ADMIN: Cleared stale selected_team_id from session")
        elif current_user.role == 'account_admin':
            account_id = current_user.account_id
            accounts = [Account.query.get(account_id)] if account_id else []
            teams = Team.query.filter_by(account_id=account_id, is_active=True).all()
            team_id = request.args.get('team_id') or session.get('selected_team_id')
            if team_id == '':
                all_teams_selected = True
            elif team_id:
                try:
                    selected_team_id = int(team_id)
                except (TypeError, ValueError):
                    selected_team_id = None
        else:
            # Regular users: use multi-team service with proper team access
            account_id = current_user.account_id
            accounts = [Account.query.get(account_id)] if account_id else []

            # Get team filter context using TeamAccessService
            team_filter_context = TeamAccessService.get_team_filter_context()
            teams = team_filter_context['user_teams']

            team_param = request.args.get('team_id')
            all_teams_selected = team_param == ''

            if team_param is None:
                # No filter provided -> default to previously selected/primary team
                selected_team_id = team_filter_context['selected_team_id']
            elif all_teams_selected:
                selected_team_id = None
            else:
                try:
                    selected_team_id = int(team_param)
                    TeamAccessService.set_selected_team(selected_team_id)
                except (TypeError, ValueError):
                    selected_team_id = team_filter_context['selected_team_id']

            team_id = selected_team_id  # maintain existing variable usage

            logger.debug(
                f"🔍 REGULAR_USER: team_param={team_param}, selected_team_id={selected_team_id}, "
                f"all_teams_selected={all_teams_selected}, user teams={len(teams)}"
            )
        # 🔧 FIX: Apply filtering logic based on user role and selections
        if current_user.role == 'super_admin':
            # For super_admin: only filter if explicitly selected
            if account_id:
                query = query.filter_by(account_id=account_id)
                logger.debug(f"🔍 SUPER_ADMIN: Filtering by account_id: {account_id}")
            else:
                logger.debug(f"🔍 SUPER_ADMIN: No account filter - showing ALL accounts")
            if team_id:
                query = query.filter_by(team_id=team_id)
                logger.debug(f"🔍 SUPER_ADMIN: Filtering by team_id: {team_id}")
            else:
                logger.debug(f"🔍 SUPER_ADMIN: No team filter - showing ALL teams")
        elif current_user.role == 'account_admin':
            # Account admin: filter by account and optional team
            if account_id:
                query = query.filter_by(account_id=account_id)
                logger.debug(f"🔍 ACCOUNT_ADMIN: Filtering by account_id: {account_id}")
            if team_id and not all_teams_selected:
                query = query.filter_by(team_id=team_id)
                logger.debug(f"🔍 ACCOUNT_ADMIN: Filtering by team_id: {team_id}")
        else:
            # Regular users: use multi-team filtering with proper account constraint
            if account_id:
                query = query.filter_by(account_id=account_id)
                logger.debug(f"🔍 REGULAR_USER: Filtering by account_id: {account_id}")
            
            # Apply team filtering using TeamAccessService
            effective_team_ids = TeamAccessService.get_effective_team_ids()
            if all_teams_selected:
                if effective_team_ids:
                    query = query.filter(Shift.team_id.in_(effective_team_ids))
                    logger.debug(f"🔍 REGULAR_USER: All Teams selected - filtering by team_ids: {effective_team_ids}")
                else:
                    query = query.filter(Shift.team_id == -1)
                    logger.debug(f"🔍 REGULAR_USER: No accessible teams")
            elif selected_team_id:
                if selected_team_id in effective_team_ids:
                    query = query.filter_by(team_id=selected_team_id)
                    logger.debug(f"🔍 REGULAR_USER: Filtering by selected team_id: {selected_team_id}")
                else:
                    query = query.filter(Shift.team_id == -1)
                    logger.debug(f"🔍 REGULAR_USER: Selected team {selected_team_id} invalid")
            else:
                # Default fallback: show primary team if available
                primary_team_id = (
                    team_filter_context['selected_team_id']
                    if team_filter_context and teams else None
                )
                if primary_team_id:
                    query = query.filter_by(team_id=primary_team_id)
                    selected_team_id = primary_team_id
                    logger.debug(f"🔍 REGULAR_USER: Fallback to primary team_id: {primary_team_id}")
                else:
                    query = query.filter(Shift.team_id == -1)
                    logger.debug(f"🔍 REGULAR_USER: No primary team available")
        if date_filter:
            try:
                date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
                query = query.filter_by(date=date_obj)  # 🔧 FIXED: Use date column
                logger.debug(f"🔍 Filtering by date: {date_obj}")
            except Exception:
                logger.debug(f"🔍 Invalid date filter: {date_filter}")
                pass
        if shift_type_filter:
            query = query.filter_by(current_shift_type=shift_type_filter)
            logger.debug(f"🔍 Filtering by shift_type: {shift_type_filter}")
        
        logger.debug(f"🔍 Final query filters applied: account_id={account_id}, team_id={team_id}, date={date_filter}, shift_type={shift_type_filter}")
        
        # Show both sent and draft handovers in reports
        query = query.filter(Shift.status.in_(['sent', 'draft']))
        logger.debug(f"🔍 Added status filter: status in ['sent', 'draft']")
        
        # Order by shift date only for now (newest first)  
        # TODO: Re-add submitted_at ordering when MySQL compatibility issue is resolved
        shifts = query.order_by(
            Shift.date.desc()
        ).all()
        
        logger.debug(f"🔍 Found {len(shifts)} shifts total")
        for shift in shifts[:3]:  # Show first 3 shifts
            logger.debug(f"🔍 Shift ID: {shift.id}, Date: {shift.date}, Type: {shift.current_shift_type}, Submitted: {shift.submitted_at}")
        
        shift_data = []
        for shift in shifts:
            incidents = Incident.query.filter_by(shift_id=shift.id).all()
            # ✨ REPORTS FIX: Show comprehensive key points for handovers
            # For DRAFT handovers: Show all relevant key points (direct + active from team)
            # For SENT handovers: Show what was active at time of submission
            
            # First get key points directly associated with this shift
            direct_key_points = ShiftKeyPoint.query.filter_by(shift_id=shift.id).all()
            
            if shift.status == 'draft':
                # DRAFT: Show all relevant active key points from the same team
                logger.debug(f"🔍 REPORTS DRAFT: Shift {shift.id} is draft, showing all active key points")
                active_key_points = ShiftKeyPoint.query.filter(
                    ShiftKeyPoint.account_id == shift.account_id,
                    ShiftKeyPoint.team_id == shift.team_id,
                    ShiftKeyPoint.status.in_(['Open', 'In Progress']),
                    ShiftKeyPoint.shift_id != shift.id  # Exclude current shift's key points to avoid duplicates
                ).all()
                
                # Combine direct key points with active ones from other shifts
                all_key_points = direct_key_points + active_key_points
                
                # Apply SAME submission filtering as dashboard first
                filtered_kps = []
                for kp in all_key_points:
                    kp_shift = Shift.query.get(kp.shift_id)
                    if kp_shift and kp_shift.status == 'sent':
                        # Check if there's a newer version that closed this key point
                        newer_closed = ShiftKeyPoint.query.filter(
                            ShiftKeyPoint.description == kp.description,
                            ShiftKeyPoint.jira_id == kp.jira_id,
                            ShiftKeyPoint.status == 'Closed',
                            ShiftKeyPoint.id > kp.id
                        ).first()
                        if newer_closed:
                            continue
                    filtered_kps.append(kp)
                
                # Apply SAME deduplication logic as dashboard
                kp_map = {}
                for kp in filtered_kps:
                    if kp.status == 'Closed':
                        continue
                    normalized_jira = kp.jira_id if kp.jira_id and kp.jira_id.lower() not in ['none', 'null', ''] else None
                    key = (kp.description, normalized_jira)
                    if key not in kp_map or kp.id > kp_map[key].id:
                        kp_map[key] = kp
                
                key_points = list(kp_map.values())
                logger.debug(f"🔍 REPORTS DRAFT: Combined {len(direct_key_points)} direct + {len(active_key_points)} active = {len(key_points)} total key points")
                
            else:
                # SENT: Show direct key points, fallback to active ones if none exist
                key_points = direct_key_points
                if len(key_points) == 0:
                    # Fallback for sent handovers with no direct key points - apply same deduplication
                    active_key_points = ShiftKeyPoint.query.filter(
                        ShiftKeyPoint.account_id == shift.account_id,
                        ShiftKeyPoint.team_id == shift.team_id,
                        ShiftKeyPoint.status.in_(['Open', 'In Progress'])
                    ).order_by(ShiftKeyPoint.id.desc()).limit(10).all()
                    
                    # Apply SAME submission filtering as dashboard first
                    filtered_kps = []
                    for kp in active_key_points:
                        kp_shift = Shift.query.get(kp.shift_id)
                        if kp_shift and kp_shift.status == 'sent':
                            # Check if there's a newer version that closed this key point
                            newer_closed = ShiftKeyPoint.query.filter(
                                ShiftKeyPoint.description == kp.description,
                                ShiftKeyPoint.jira_id == kp.jira_id,
                                ShiftKeyPoint.status == 'Closed',
                                ShiftKeyPoint.id > kp.id
                            ).first()
                            if newer_closed:
                                continue
                        filtered_kps.append(kp)
                    
                    # Apply SAME deduplication logic as dashboard
                    kp_map = {}
                    for kp in filtered_kps:
                        if kp.status == 'Closed':
                            continue
                        normalized_jira = kp.jira_id if kp.jira_id and kp.jira_id.lower() not in ['none', 'null', ''] else None
                        key = (kp.description, normalized_jira)
                        if key not in kp_map or kp.id > kp_map[key].id:
                            kp_map[key] = kp
                    
                    key_points = list(kp_map.values())
                    logger.debug(f"🔍 REPORTS SENT: Using {len(active_key_points)} active key points (deduplicated to {len(key_points)}) as fallback")
            
            logger.debug(f"🔍 REPORTS SHIFT {shift.id}: Found {len(incidents)} incidents, {len(key_points)} key points")
            
            # Debug key points details with status breakdown
            status_counts = {'Open': 0, 'In Progress': 0, 'Closed': 0}
            for kp in key_points:
                status_counts[kp.status] = status_counts.get(kp.status, 0) + 1
                logger.debug(f"  📋 REPORTS KP {kp.id}: '{kp.description[:30]}...' - Status: {kp.status} - Responsible: {kp.responsible_engineer_id}")
            
            logger.debug(f"  📊 REPORTS Status breakdown for shift {shift.id}: {status_counts}")
            
            if len(key_points) == 0:
                logger.debug(f"  ⚠️  REPORTS WARNING: No key points found for shift {shift.id}")
            
            # Get detailed incident information
            incidents_data = []
            for inc in incidents:
                incident_details = {
                    'type': inc.type,
                    'title': inc.title,
                    'status': inc.status,
                    'priority': inc.priority,
                    'handover': inc.handover,
                    'assigned_to': inc.assigned_to,
                    'description': inc.description,
                    'escalated_to': inc.escalated_to
                }
                # Debug logging for incident assignment tracking
                if inc.assigned_to:
                    logger.debug(f"🔍 Incident {inc.id} has assignment: '{inc.assigned_to}'")
                incidents_data.append(incident_details)
            
            # Get detailed key points information - INCLUDING ALL STATUSES for reports
            key_points_data = []
            for kp in key_points:
                try:
                    engineer = None
                    if kp.responsible_engineer_id:
                        engineer = TeamMember.query.get(kp.responsible_engineer_id)
                    
                    kp_data = {
                        'description': kp.description,
                        'status': kp.status,
                        'responsible': engineer.name if engineer else 'N/A',
                        'jira_id': kp.jira_id,
                        'id': kp.id  # Add ID for debugging
                    }
                    key_points_data.append(kp_data)
                    logger.debug(f"  ✅ REPORTS Added KP {kp.id} to template data: Status={kp.status}")
                    
                except Exception as e:
                    logger.warning(f"🚨 Error processing key point {kp.id}: {e}")
                    # Add basic key point without responsible engineer
                    kp_data = {
                        'description': kp.description,
                        'status': kp.status,
                        'responsible': 'Error',
                        'jira_id': kp.jira_id,
                        'id': kp.id
                    }
                    key_points_data.append(kp_data)
                    logger.debug(f"  ⚠️  REPORTS Added KP {kp.id} with error to template data")
            
            # 🔧 FIX: Find who submitted this handover from HandoverRequest table (more accurate)
            submitted_by = 'Unknown'
            try:
                # Import here to avoid circular imports
                from models.handover_enhanced import HandoverRequest
                
                # Find the corresponding handover request with exact matching
                handover_req = HandoverRequest.query.filter_by(
                    shift_date=shift.date,
                    current_shift_type=shift.current_shift_type,
                    account_id=shift.account_id,
                    team_id=shift.team_id
                ).first()
                
                if handover_req and handover_req.created_by_id:
                    user = User.query.get(handover_req.created_by_id)
                    if user:
                        submitted_by = user.display_name or user.username
                        logger.debug(f"🔍 REPORTS: Found accurate submitter: {submitted_by} (ID: {user.id}) for shift {shift.id}")
                    else:
                        submitted_by = f'User ID: {handover_req.created_by_id}'
                        logger.debug(f"🔍 REPORTS: User not found for ID: {handover_req.created_by_id}")
                else:
                    logger.debug(f"🔍 REPORTS: No HandoverRequest found for shift {shift.id}, trying audit log fallback...")
                    # Fallback to audit log with more specific matching
                    from models.audit_log import AuditLog
                    audit_entry = AuditLog.query.filter(
                        AuditLog.action.like('%Create Handover%'),
                        AuditLog.details.like(f'%Team: {shift.team_id}%'),
                        AuditLog.details.like(f'%Account: {shift.account_id}%'),
                        AuditLog.details.like(f'%Date: {shift.date}%')
                    ).first()
                    
                    if audit_entry and audit_entry.user_id:
                        user = User.query.get(audit_entry.user_id)
                        if user:
                            submitted_by = user.display_name or user.username
                            logger.debug(f"🔍 REPORTS: Fallback audit submitter: {submitted_by} for shift {shift.id}")
                        else:
                            submitted_by = audit_entry.username or 'Unknown User'
                    
            except Exception as e:
                logger.debug(f"🚨 Error finding submitter for shift {shift.id}: {e}")
                submitted_by = 'Unknown'
            
            # Get Change Info and KB Updates for this date/team with deduplication
            from datetime import timedelta
            raw_change_infos = ShiftChangeInfo.query.filter(
                ShiftChangeInfo.account_id == shift.account_id,
                ShiftChangeInfo.team_id == shift.team_id,
                ShiftChangeInfo.created_at >= shift.date,
                ShiftChangeInfo.created_at < shift.date + timedelta(days=1),
                ~ShiftChangeInfo.status.in_(['Completed', 'Cancelled', 'Implemented'])
            ).order_by(ShiftChangeInfo.id.desc()).all()
            # Deduplicate by change_number
            change_map = {}
            for ci in raw_change_infos:
                key = ci.change_number.strip().lower() if ci.change_number else f"{ci.app_name}_{ci.description[:30] if ci.description else ''}"
                if key not in change_map:
                    change_map[key] = ci
            change_infos = list(change_map.values())
            change_infos_data = []
            for change_info in change_infos:
                engineer = None
                if change_info.responsible_engineer_id:
                    engineer = TeamMember.query.get(change_info.responsible_engineer_id)
                
                change_infos_data.append({
                    'app_name': change_info.app_name,
                    'change_number': change_info.change_number,
                    'description': change_info.description,
                    'change_datetime': change_info.change_datetime,
                    'responsible': engineer.name if engineer else 'N/A',
                    'status': change_info.status  # Include status for display in reports
                })
            
            # 🔧 FIX: Query KB updates by date/team/account with deduplication
            raw_kb_updates = ShiftKBUpdate.query.filter(
                ShiftKBUpdate.account_id == shift.account_id,
                ShiftKBUpdate.team_id == shift.team_id,
                ShiftKBUpdate.created_at >= shift.date,
                ShiftKBUpdate.created_at < shift.date + timedelta(days=1),
                ShiftKBUpdate.status != 'Published'
            ).order_by(ShiftKBUpdate.id.desc()).all()
            # Deduplicate by kb_number
            kb_map = {}
            for kb in raw_kb_updates:
                key = kb.kb_number.strip().lower() if kb.kb_number else f"{kb.app_name}_{kb.description[:30] if kb.description else ''}"
                if key not in kb_map:
                    kb_map[key] = kb
            kb_updates = list(kb_map.values())
            kb_updates_data = []
            for kb_update in kb_updates:
                engineer = None
                if kb_update.responsible_engineer_id:
                    engineer = TeamMember.query.get(kb_update.responsible_engineer_id)
                
                kb_updates_data.append({
                    'app_name': kb_update.app_name,
                    'kb_number': kb_update.kb_number,
                    'description': kb_update.description,
                    'status': kb_update.status,
                    'responsible': engineer.name if engineer else 'N/A'
                })

            # Create serializable shift data instead of passing the model object
            shift_data.append({
                'shift': {
                    'id': shift.id,
                    'date': shift.date,
                    'current_shift_type': shift.current_shift_type,
                    'next_shift_type': shift.next_shift_type,
                    'status': shift.status,
                    'submitted_at': shift.submitted_at,
                    'additional_notes': shift.additional_notes,
                    'account_id': shift.account_id,
                    'team_id': shift.team_id
                },
                'incidents': incidents_data,
                'key_points': key_points_data,
                'change_infos': change_infos_data,
                'kb_updates': kb_updates_data,
                'submitted_by': submitted_by
            })
            
            # Debug what we're sending to template
            logger.debug(f"  📋 REPORTS Sending {len(key_points_data)} key points, {len(change_infos_data)} change_infos, {len(kb_updates_data)} kb_updates to template for shift {shift.id}")
            if len(change_infos_data) > 0 or len(kb_updates_data) > 0:
                logger.debug(f"  🔥 SHIFT {shift.id} HAS CHANGE/KB DATA - Should show extra sections!")
            for kp_data in key_points_data:
                logger.debug(f"    - KP {kp_data.get('id', 'Unknown')}: {kp_data['status']} - {kp_data['description'][:30]}...")
        
        # Calculate visualization data
        total_shifts = len(shift_data)
        total_incidents = sum(len(entry['incidents']) for entry in shift_data)
        total_keypoints = sum(len(entry['key_points']) for entry in shift_data)
        
        # Incident type distribution
        incident_types = {}
        incident_priorities = {'High': 0, 'Medium': 0, 'Low': 0}
        keypoint_statuses = {'Open': 0, 'In Progress': 0, 'Closed': 0}
        
        shift_type_distribution = {'Morning': 0, 'Evening': 0, 'Night': 0}
        
        for entry in shift_data:
            # Shift type distribution
            shift_type = entry['shift']['current_shift_type']
            if shift_type in shift_type_distribution:
                shift_type_distribution[shift_type] += 1
                
            # Incident analysis
            for inc in entry['incidents']:
                inc_type = inc['type']
                incident_types[inc_type] = incident_types.get(inc_type, 0) + 1
                
                priority = inc['priority']
                if priority in incident_priorities:
                    incident_priorities[priority] += 1
            
            # Key point analysis
            for kp in entry['key_points']:
                status = kp['status']
                if status in keypoint_statuses:
                    keypoint_statuses[status] += 1
        
        stats = {
            'total_shifts': total_shifts,
            'total_incidents': total_incidents,
            'total_keypoints': total_keypoints,
            'incident_types': incident_types,
            'incident_priorities': incident_priorities,
            'keypoint_statuses': keypoint_statuses,
            'shift_type_distribution': shift_type_distribution
        }
        
        # 🔍 DEBUG: Final data being sent to template
        logger.debug(f"🔍 REPORTS DEBUG: Sending {len(shift_data)} shifts to template")
        logger.debug(f"🔍 REPORTS DEBUG: Key Point Status Summary: {keypoint_statuses}")
        logger.debug(f"🔍 REPORTS DEBUG: Stats = {stats}")
        
        # Verify that closed key points are included in the data
        total_closed_in_data = 0
        for data in shift_data:
            closed_count = sum(1 for kp in data['key_points'] if kp['status'] == 'Closed')
            total_closed_in_data += closed_count
            if closed_count > 0:
                logger.debug(f"🔍 REPORTS: Shift {data['shift']['id']} has {closed_count} closed key points in template data")
        
        logger.debug(f"🔍 REPORTS FINAL: Total closed key points being sent to template: {total_closed_in_data}")
        
        # 🔧 FIX: Pre-sort and pre-group shift data by date for template
        if shift_data:
            shift_data.sort(key=lambda x: x['shift']['date'], reverse=True)
            logger.debug(f"🔍 REPORTS SORT: Sorted {len(shift_data)} shifts by date")
            for i, data in enumerate(shift_data[:5]):
                logger.debug(f"  {i+1}. {data['shift']['date']} ({data['shift']['current_shift_type']})")
            
            # Pre-group shifts by date in the correct order
            grouped_shift_data = OrderedDict()
            for entry in shift_data:
                date_str = entry['shift']['date'].strftime('%Y-%m-%d')
                if date_str not in grouped_shift_data:
                    grouped_shift_data[date_str] = []
                grouped_shift_data[date_str].append(entry)
            
            logger.debug(f"🔍 REPORTS GROUPS: Created {len(grouped_shift_data)} date groups")
            for date_str, entries in list(grouped_shift_data.items())[:3]:
                logger.debug(f"  {date_str}: {len(entries)} shifts")
        else:
            grouped_shift_data = OrderedDict()
        
        if shift_data:
            logger.debug(f"🔍 REPORTS DEBUG: First 3 shifts:")
            for i, data in enumerate(shift_data[:3]):
                shift = data['shift']
                closed_kps = [kp for kp in data['key_points'] if kp['status'] == 'Closed']
                logger.debug(f"🔍   Shift {i+1}: ID={shift['id']}, Date={shift['date']}, Type={shift['current_shift_type']}, Status={shift['status']}, Incidents={len(data['incidents'])}, KeyPoints={len(data['key_points'])}, Closed KPs={len(closed_kps)}")
        else:
            logger.debug(f"🔍 REPORTS DEBUG: No shift data to display!")
        
        # Extract ordered date keys to pass to template for explicit ordering
        ordered_date_keys = list(grouped_shift_data.keys()) if grouped_shift_data else []
        logger.debug(f"🔍 REPORTS ORDERED KEYS: {ordered_date_keys}")
        
        # Add cache busting to ensure fresh data
        import time
        cache_buster = int(time.time())
        
        # Get team filter context for multi-team users
        if current_user.role not in ['super_admin', 'account_admin']:
            team_filter_context = TeamAccessService.get_team_filter_context()
        else:
            team_filter_context = None
        
        return render_template(
            'handover_reports.html',
            shift_data=shift_data,
            grouped_shift_data=grouped_shift_data,
            ordered_date_keys=ordered_date_keys,
            stats=stats,
            date_filter=date_filter or '',
            shift_type_filter=shift_type_filter or '',
            accounts=accounts,
            teams=teams,
            selected_account_id=account_id,
            selected_team_id=selected_team_id,
            cache_buster=cache_buster,
            team_filter_context=team_filter_context
        )
    except Exception as e:
        logger.error(f"❌ Error in handover_reports: {e}")
        import traceback
        traceback.print_exc()
        return render_template(
            'handover_reports.html',
            shift_data=[],
            stats={},
            error=f"Error loading reports: {str(e)}",
            date_filter='',
            shift_type_filter='',
            accounts=[],
            teams=[],
            selected_account_id=None,
            selected_team_id=None,
            team_filter_context=None
        )


# API endpoints for inline editing
@reports_bp.route('/api/kb-updates/<int:kb_id>', methods=['PUT'])
@login_required
def update_kb_record(kb_id):
    """Update a KB update record"""
    try:
        kb_update = ShiftKBUpdate.query.get_or_404(kb_id)
        
        # Check permissions
        if current_user.role not in ['super_admin', 'account_admin']:
            if kb_update.account_id != current_user.account_id or kb_update.team_id != current_user.team_id:
                return jsonify({'error': 'Unauthorized'}), 403
        elif current_user.role == 'account_admin':
            if kb_update.account_id != current_user.account_id:
                return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        
        # Update fields
        kb_update.app_name = data.get('app_name', kb_update.app_name)
        kb_update.kb_number = data.get('kb_number', kb_update.kb_number)
        kb_update.description = data.get('description', kb_update.description)
        kb_update.status = data.get('status', kb_update.status)
        
        # Handle responsible engineer
        responsible_engineer_id = data.get('responsible_engineer_id')
        if responsible_engineer_id:
            kb_update.responsible_engineer_id = responsible_engineer_id
        else:
            kb_update.responsible_engineer_id = None
        
        db.session.commit()
        
        # Get responsible engineer name for response
        responsible_engineer_name = None
        if kb_update.responsible_engineer:
            responsible_engineer_name = kb_update.responsible_engineer.name
        
        log_action('Update KB Record', f'Updated KB record {kb_id}')
        
        return jsonify({
            'success': True,
            'message': 'KB update saved successfully',
            'responsible_engineer_name': responsible_engineer_name
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/api/kb-updates/<int:kb_id>', methods=['DELETE'])
@login_required
def delete_kb_record(kb_id):
    """Delete a KB update record"""
    try:
        kb_update = ShiftKBUpdate.query.get_or_404(kb_id)
        
        # Check permissions - only admins and team members can delete
        if current_user.role not in ['super_admin', 'account_admin', 'team_admin']:
            if kb_update.account_id != current_user.account_id or kb_update.team_id != current_user.team_id:
                return jsonify({'error': 'Unauthorized - insufficient permissions'}), 403
        elif current_user.role == 'account_admin':
            if kb_update.account_id != current_user.account_id:
                return jsonify({'error': 'Unauthorized - different account'}), 403
        elif current_user.role == 'team_admin':
            if kb_update.account_id != current_user.account_id or kb_update.team_id != current_user.team_id:
                return jsonify({'error': 'Unauthorized - different team'}), 403
        
        kb_number = kb_update.kb_number or 'N/A'
        app_name = kb_update.app_name or 'N/A'
        
        db.session.delete(kb_update)
        db.session.commit()
        
        log_action('Delete KB Record', f'Deleted KB record {kb_id}: {app_name} - {kb_number}')
        
        return jsonify({
            'success': True,
            'message': f'KB record deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting KB record {kb_id}: {str(e)}")
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/api/team-members', methods=['GET'])
@login_required
def get_team_members():
    """Get team members for dropdown"""
    try:
        logger.debug(f"[DEBUG] Getting team members for user {current_user.id} role {current_user.role}")
        query = TeamMember.query
        
        # Filter by user permissions
        if current_user.role == 'super_admin':
            logger.debug("[DEBUG] Super admin - getting all team members")
            pass  # Can see all team members
        elif current_user.role == 'account_admin':
            logger.debug(f"[DEBUG] Account admin - filtering by account {current_user.account_id}")
            query = query.filter_by(account_id=current_user.account_id)
        else:
            logger.debug(f"[DEBUG] Regular user - filtering by account {current_user.account_id} and team {current_user.team_id}")
            query = query.filter_by(account_id=current_user.account_id, team_id=current_user.team_id)
        
        members = query.all()
        logger.debug(f"[DEBUG] Found {len(members)} team members")
        
        result = [{
            'id': member.id,
            'name': member.name
        } for member in members]
        
        logger.debug(f"[DEBUG] Returning team members: {result}")
        return jsonify(result)
        
    except Exception as e:
        logger.debug(f"[DEBUG] Error getting team members: {str(e)}")
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/api/change-info/<int:change_id>', methods=['PUT'])
@login_required
def update_change_info_record(change_id):
    """Update a change info record"""
    try:
        change_info = ShiftChangeInfo.query.get_or_404(change_id)
        
        # Check permissions
        if current_user.role not in ['super_admin', 'account_admin']:
            if change_info.account_id != current_user.account_id or change_info.team_id != current_user.team_id:
                return jsonify({'error': 'Unauthorized'}), 403
        elif current_user.role == 'account_admin':
            if change_info.account_id != current_user.account_id:
                return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        
        # Update fields
        change_info.app_name = data.get('app_name', change_info.app_name)
        change_info.change_number = data.get('change_number', change_info.change_number)
        change_info.description = data.get('description', change_info.description)
        change_info.status = data.get('status', change_info.status)
        
        # Handle change datetime
        change_datetime_str = data.get('change_datetime')
        if change_datetime_str:
            try:
                change_info.change_datetime = datetime.fromisoformat(change_datetime_str.replace('Z', '+00:00'))
            except ValueError:
                change_info.change_datetime = None
        else:
            change_info.change_datetime = None
        
        # Handle responsible engineer
        responsible_engineer_id = data.get('responsible_engineer_id')
        if responsible_engineer_id:
            change_info.responsible_engineer_id = responsible_engineer_id
        else:
            change_info.responsible_engineer_id = None
        
        db.session.commit()
        
        # Get responsible engineer name for response
        responsible_engineer_name = None
        if change_info.responsible_engineer:
            responsible_engineer_name = change_info.responsible_engineer.name
        
        log_action('Update Change Info', f'Updated change info record {change_id}')
        
        return jsonify({
            'success': True,
            'message': 'Change info saved successfully',
            'responsible_engineer_name': responsible_engineer_name
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/api/change-info/<int:change_id>', methods=['DELETE'])
@login_required
def delete_change_info_record(change_id):
    """Delete a change info record"""
    try:
        change_info = ShiftChangeInfo.query.get_or_404(change_id)
        
        # Check permissions - only admins and team members can delete
        if current_user.role not in ['super_admin', 'account_admin', 'team_admin']:
            if change_info.account_id != current_user.account_id or change_info.team_id != current_user.team_id:
                return jsonify({'error': 'Unauthorized - insufficient permissions'}), 403
        elif current_user.role == 'account_admin':
            if change_info.account_id != current_user.account_id:
                return jsonify({'error': 'Unauthorized - different account'}), 403
        elif current_user.role == 'team_admin':
            if change_info.account_id != current_user.account_id or change_info.team_id != current_user.team_id:
                return jsonify({'error': 'Unauthorized - different team'}), 403
        
        change_number = change_info.change_number or 'N/A'
        app_name = change_info.app_name or 'N/A'
        
        db.session.delete(change_info)
        db.session.commit()
        
        log_action('Delete Change Info', f'Deleted change info record {change_id}: {app_name} - {change_number}')
        
        return jsonify({
            'success': True,
            'message': f'Change info record deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting change info record {change_id}: {str(e)}")
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/api/change-info', methods=['POST'])
@login_required
def create_change_info_record():
    """Create a new change info record"""
    try:
        data = request.get_json()
        
        # Get current shift or create one for today
        from datetime import date
        today = date.today()
        current_shift = Shift.query.filter_by(
            date=today,
            account_id=current_user.account_id,
            team_id=current_user.team_id
        ).first()
        
        if not current_shift:
            # Create a new shift for today
            current_shift = Shift(
                date=today,
                current_shift_type='Morning',  # Default
                next_shift_type='Evening',     # Default
                account_id=current_user.account_id,
                team_id=current_user.team_id
            )
            db.session.add(current_shift)
            db.session.flush()  # Get the ID
        
        # Create new change info record
        change_info = ShiftChangeInfo(
            app_name=data.get('app_name'),
            change_number=data.get('change_number'),
            description=data.get('description'),
            status=data.get('status', 'New'),
            shift_id=current_shift.id,
            account_id=current_user.account_id,
            team_id=current_user.team_id
        )
        
        # Handle change datetime
        change_datetime_str = data.get('change_datetime')
        if change_datetime_str:
            try:
                change_info.change_datetime = datetime.fromisoformat(change_datetime_str.replace('Z', '+00:00'))
            except ValueError:
                change_info.change_datetime = None
        
        # Handle responsible engineer
        responsible_engineer_id = data.get('responsible_engineer_id')
        if responsible_engineer_id:
            change_info.responsible_engineer_id = responsible_engineer_id
        
        db.session.add(change_info)
        db.session.commit()
        
        log_action('Create Change Info', f'Created new change info record: {change_info.change_number}')
        
        return jsonify({
            'success': True,
            'message': 'Change info created successfully',
            'id': change_info.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/api/kb-updates', methods=['POST'])
@login_required
def create_kb_update_record():
    """Create a new KB update record"""
    try:
        data = request.get_json()
        
        # Get current shift or create one for today
        from datetime import date
        today = date.today()
        current_shift = Shift.query.filter_by(
            date=today,
            account_id=current_user.account_id,
            team_id=current_user.team_id
        ).first()
        
        if not current_shift:
            # Create a new shift for today
            current_shift = Shift(
                date=today,
                current_shift_type='Morning',  # Default
                next_shift_type='Evening',     # Default
                account_id=current_user.account_id,
                team_id=current_user.team_id
            )
            db.session.add(current_shift)
            db.session.flush()  # Get the ID
        
        # Create new KB update record
        kb_update = ShiftKBUpdate(
            app_name=data.get('app_name'),
            kb_number=data.get('kb_number'),
            description=data.get('description'),
            status=data.get('status', 'New'),
            shift_id=current_shift.id,
            account_id=current_user.account_id,
            team_id=current_user.team_id
        )
        
        # Handle responsible engineer
        responsible_engineer_id = data.get('responsible_engineer_id')
        if responsible_engineer_id:
            kb_update.responsible_engineer_id = responsible_engineer_id
        
        db.session.add(kb_update)
        db.session.commit()
        
        log_action('Create KB Update', f'Created new KB update record: {kb_update.kb_number}')
        
        return jsonify({
            'success': True,
            'message': 'KB update created successfully',
            'id': kb_update.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/handover-reports/delete-draft/<int:shift_id>', methods=['POST'])
@login_required
def delete_draft_report(shift_id):
    """Delete a draft shift report - Super Admin only"""
    try:
        # Check if user is super_admin
        if current_user.role != 'super_admin':
            flash('Access denied. Only Super Admins can delete draft reports.', 'error')
            return redirect(url_for('reports.handover_reports'))
        
        # Get the shift
        shift = Shift.query.get(shift_id)
        
        if not shift:
            flash('Shift report not found.', 'error')
            return redirect(url_for('reports.handover_reports'))
        
        # Check if it's a draft
        if shift.status != 'draft':
            flash('Only draft reports can be deleted. Submitted reports cannot be deleted.', 'error')
            return redirect(url_for('reports.handover_reports'))
        
        # Store info for logging
        shift_date = shift.date.strftime('%Y-%m-%d') if shift.date else 'Unknown'
        shift_type = f"{shift.current_shift_type} → {shift.next_shift_type}"
        team_name = shift.team.name if shift.team else 'Unknown'
        
        # Delete related records first
        # Delete incidents
        deleted_incidents = Incident.query.filter_by(shift_id=shift_id).delete()
        
        # Delete key points
        deleted_keypoints = ShiftKeyPoint.query.filter_by(shift_id=shift_id).delete()
        
        # Delete change info
        deleted_changes = ShiftChangeInfo.query.filter_by(shift_id=shift_id).delete()
        
        # Delete KB updates
        deleted_kbs = ShiftKBUpdate.query.filter_by(shift_id=shift_id).delete()
        
        # Delete the shift itself
        db.session.delete(shift)
        db.session.commit()
        
        # Log the action
        log_action('Delete Draft Report', 
                  f'Deleted draft report: Date={shift_date}, Shift={shift_type}, Team={team_name}. '
                  f'Deleted: {deleted_incidents} incidents, {deleted_keypoints} key points, '
                  f'{deleted_changes} changes, {deleted_kbs} KB updates.')
        
        flash(f'Draft report for {shift_date} ({shift_type}) has been deleted successfully.', 'success')
        logger.info(f"🗑️ Super Admin {current_user.username} deleted draft report ID {shift_id}")
        
        return redirect(url_for('reports.handover_reports'))
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Error deleting draft report {shift_id}: {e}")
        flash(f'Error deleting draft report: {str(e)}', 'error')
        return redirect(url_for('reports.handover_reports'))


@reports_bp.route('/api/delete-draft/<int:shift_id>', methods=['DELETE'])
@login_required
def api_delete_draft_report(shift_id):
    """API endpoint to delete a draft shift report - Super Admin only"""
    try:
        # Check if user is super_admin
        if current_user.role != 'super_admin':
            return jsonify({'success': False, 'error': 'Access denied. Only Super Admins can delete draft reports.'}), 403
        
        # Get the shift
        shift = Shift.query.get(shift_id)
        
        if not shift:
            return jsonify({'success': False, 'error': 'Shift report not found.'}), 404
        
        # Check if it's a draft
        if shift.status != 'draft':
            return jsonify({'success': False, 'error': 'Only draft reports can be deleted.'}), 400
        
        # Store info for logging
        shift_date = shift.date.strftime('%Y-%m-%d') if shift.date else 'Unknown'
        shift_type = f"{shift.current_shift_type} → {shift.next_shift_type}"
        team_name = shift.team.name if shift.team else 'Unknown'
        
        # Delete related records first
        deleted_incidents = Incident.query.filter_by(shift_id=shift_id).delete()
        deleted_keypoints = ShiftKeyPoint.query.filter_by(shift_id=shift_id).delete()
        deleted_changes = ShiftChangeInfo.query.filter_by(shift_id=shift_id).delete()
        deleted_kbs = ShiftKBUpdate.query.filter_by(shift_id=shift_id).delete()
        
        # Delete the shift itself
        db.session.delete(shift)
        db.session.commit()
        
        # Log the action
        log_action('Delete Draft Report', 
                  f'Deleted draft report: Date={shift_date}, Shift={shift_type}, Team={team_name}')
        
        logger.info(f"🗑️ Super Admin {current_user.username} deleted draft report ID {shift_id} via API")
        
        return jsonify({
            'success': True,
            'message': f'Draft report for {shift_date} ({shift_type}) deleted successfully.',
            'deleted': {
                'incidents': deleted_incidents,
                'keypoints': deleted_keypoints,
                'changes': deleted_changes,
                'kb_updates': deleted_kbs
            }
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Error deleting draft report {shift_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
