"""
Shift Allowance Report Generation Module

This module generates Excel reports for shift allowance tracking.
The report contains two sheets:
- SC RD's: Report Directors (specific members identified by name)
- SC Team: All other team members

Shift Mapping Rules:
- For ALL engineers: N (Night) and LE (Late Evening) → Night Shift
- For SC RD's ONLY: E (Evening) → UK
- All other shift codes → Empty (blank)
"""

import logging
from flask import Blueprint, request, send_file, flash, redirect, url_for, jsonify
from flask_login import login_required, current_user
from models.models import db, TeamMember, ShiftRoster, Account, Team
from datetime import datetime, date
from io import BytesIO
import calendar

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

shift_allowance_bp = Blueprint('shift_allowance', __name__)

# ========================================
# CONFIGURATION - Easily modifiable
# ========================================

# List of RD (Report Director) engineer names
# These engineers have special shift mapping rules (E → UK)
RD_ENGINEER_NAMES = [
    "Garima Singh",
    "Atharva Tiwari",
    "Shodha Hm",
    "Shodha HM",  # Handle case variations
]

# Shift code mappings for allowance calculation
# Maps roster shift codes to allowance report display values
SHIFT_MAPPINGS = {
    # Night shifts - applicable to ALL engineers
    'N': 'Night',      # Night shift → Display as "Night"
    'LE': 'Night',     # Late Evening → Treated as Night shift
    
    # For RD engineers ONLY
    'E': 'UK',         # Evening → UK (only for RDs)
}

# Shift codes that qualify for night shift allowance (for ALL engineers)
NIGHT_SHIFT_CODES = ['N', 'LE']

# Shift code for UK allowance (RD engineers only)
UK_SHIFT_CODE = 'E'

# ========================================
# HELPER FUNCTIONS
# ========================================

def is_rd_engineer(member_name):
    """
    Check if a team member is an RD (Report Director) engineer.
    RD engineers have special shift mapping rules.
    
    Args:
        member_name: Name of the team member
        
    Returns:
        bool: True if member is an RD engineer
    """
    if not member_name:
        return False
    
    # Normalize name for comparison (case-insensitive, strip whitespace)
    normalized_name = member_name.strip().lower()
    
    for rd_name in RD_ENGINEER_NAMES:
        if rd_name.strip().lower() == normalized_name:
            return True
    return False


def get_allowance_value(shift_code, is_rd=False):
    """
    Get the shift allowance display value based on shift code and engineer type.
    
    Business Rules:
    - For ALL engineers: N and LE → "Night"
    - For RD engineers ONLY: E → "UK"
    - All other codes → Empty string (blank)
    
    Args:
        shift_code: The roster shift code (e.g., 'N', 'E', 'D', 'LE')
        is_rd: Whether the engineer is an RD (Report Director)
        
    Returns:
        str: Display value for the allowance report
    """
    if not shift_code:
        return ''
    
    shift_code = str(shift_code).strip().upper()
    
    # Night shifts apply to ALL engineers
    if shift_code in NIGHT_SHIFT_CODES:
        return 'Night'
    
    # UK (Evening) only applies to RD engineers
    if is_rd and shift_code == UK_SHIFT_CODE:
        return 'UK'
    
    # All other cases - return empty
    return ''


def get_month_dates(year, month):
    """
    Get all dates for a given month.
    
    Args:
        year: Year (e.g., 2025)
        month: Month (1-12)
        
    Returns:
        list: List of date objects for the month
    """
    num_days = calendar.monthrange(year, month)[1]
    return [date(year, month, day) for day in range(1, num_days + 1)]


def generate_shift_allowance_excel(account_id, team_id, year, month, 
                                    rd_member_ids=None, rd_member_names=None, 
                                    exception_member_ids=None):
    """
    Generate Excel report with shift allowance data.
    
    Creates a workbook with two sheets:
    - SC RD's: Report Directors with special UK mapping
    - SC Team: All other team members
    
    Args:
        account_id: Account ID to filter data
        team_id: Team ID to filter data
        year: Year for the report
        month: Month for the report (1-12)
        rd_member_ids: List of member IDs to include as RDs (if provided, overrides name-based detection)
        rd_member_names: List of member names to include as RDs (used with rd_member_ids)
        exception_member_ids: List of member IDs to exclude from the report entirely
        
    Returns:
        BytesIO: Excel file as bytes buffer
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Please install it: pip install openpyxl")
        raise ImportError("openpyxl is required for Excel generation")
    
    # Initialize lists if None
    rd_member_ids = rd_member_ids or []
    rd_member_names = rd_member_names or []
    exception_member_ids = exception_member_ids or []
    
    # Get all dates for the month
    month_dates = get_month_dates(year, month)
    month_name = calendar.month_name[month]
    
    # Query team members for the specified account/team
    members_query = TeamMember.query.filter_by(
        account_id=account_id,
        team_id=team_id,
        is_active=True
    ).order_by(TeamMember.name)
    
    all_members = members_query.all()
    
    # Filter out exception members
    if exception_member_ids:
        all_members = [m for m in all_members if m.id not in exception_member_ids]
        logger.info(f"[SHIFT ALLOWANCE] Excluded {len(exception_member_ids)} exception members")
    
    # Separate RD members and regular team members
    rd_members = []
    team_members = []
    
    # Use provided rd_member_ids if available, otherwise fall back to name-based detection
    if rd_member_ids:
        # Use explicitly selected RD members
        for member in all_members:
            if member.id in rd_member_ids:
                rd_members.append(member)
            else:
                team_members.append(member)
        logger.info(f"[SHIFT ALLOWANCE] Using custom RD selection: {len(rd_member_ids)} RDs selected")
    else:
        # Fall back to name-based detection (original behavior)
        for member in all_members:
            if is_rd_engineer(member.name):
                rd_members.append(member)
            else:
                team_members.append(member)
    
    logger.info(f"[SHIFT ALLOWANCE] Found {len(rd_members)} RD members, {len(team_members)} team members")
    
    # Get roster data for all members
    roster_query = ShiftRoster.query.filter(
        ShiftRoster.account_id == account_id,
        ShiftRoster.team_id == team_id,
        db.extract('month', ShiftRoster.date) == month,
        db.extract('year', ShiftRoster.date) == year
    )
    
    roster_entries = roster_query.all()
    
    # Build roster lookup: {member_id: {date: shift_code}}
    roster_data = {}
    for entry in roster_entries:
        if entry.team_member_id not in roster_data:
            roster_data[entry.team_member_id] = {}
        roster_data[entry.team_member_id][entry.date] = entry.shift_code
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF00B050', end_color='FF00B050', fill_type='solid')  # Green
    header_fill_pink = PatternFill(start_color='FFFF00FF', end_color='FFFF00FF', fill_type='solid')  # Magenta/Pink
    night_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')  # Yellow for Night
    uk_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')  # Yellow for UK
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    def create_sheet(ws, members, is_rd_sheet=False):
        """Create a sheet with member data"""
        
        # Row 1: Header row
        headers = ['Name', 'UID', 'Project']
        
        # Add date columns (format: day-month, e.g., "1-Nov", "15-Nov")
        for d in month_dates:
            # Cross-platform date formatting (%-d doesn't work on Windows)
            headers.append(f"{d.day}-{d.strftime('%b')}")
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_pink if col_idx <= 3 else header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Write member data
        for row_idx, member in enumerate(members, 2):
            # Name column
            name_cell = ws.cell(row=row_idx, column=1, value=member.name)
            name_cell.alignment = left_align
            name_cell.border = thin_border
            
            # Employee ID / UID column
            uid_cell = ws.cell(row=row_idx, column=2, value=member.employee_id or '')
            uid_cell.alignment = center_align
            uid_cell.border = thin_border
            
            # Project column (can be customized - using team name or static value)
            project_cell = ws.cell(row=row_idx, column=3, value='CTCO-OPSC')
            project_cell.alignment = center_align
            project_cell.border = thin_border
            
            # Date columns with shift allowance values
            member_roster = roster_data.get(member.id, {})
            
            for col_offset, d in enumerate(month_dates):
                col_idx = col_offset + 4  # Starting from column 4
                shift_code = member_roster.get(d, '')
                
                # Get allowance value based on shift code and member type
                allowance_value = get_allowance_value(shift_code, is_rd=is_rd_sheet)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=allowance_value)
                cell.alignment = center_align
                cell.border = thin_border
                
                # Apply yellow highlighting for Night and UK values
                if allowance_value in ['Night', 'UK']:
                    cell.fill = night_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25  # Name column
        ws.column_dimensions['B'].width = 10  # UID column
        ws.column_dimensions['C'].width = 12  # Project column
        
        for col_idx in range(4, 4 + len(month_dates)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 8
    
    # Create Sheet 1: SC RD's
    ws_rds = wb.active
    ws_rds.title = "SC RD's"
    create_sheet(ws_rds, rd_members, is_rd_sheet=True)
    
    # Create Sheet 2: SC Team
    ws_team = wb.create_sheet(title="SC Team")
    create_sheet(ws_team, team_members, is_rd_sheet=False)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info(f"[SHIFT ALLOWANCE] Generated report for {month_name} {year}")
    
    return output


# ========================================
# ROUTES
# ========================================

@shift_allowance_bp.route('/api/shift-allowance/download', methods=['GET', 'POST'])
@login_required
def download_shift_allowance():
    """
    Download shift allowance Excel report.
    
    POST JSON Body (preferred):
        - month: Month number (1-12)
        - year: Year (e.g., 2025)
        - account_id: Account ID (optional for super_admin)
        - team_id: Team ID
        - rd_member_ids: List of member IDs to include in RD sheet
        - rd_member_names: List of member names for RD sheet
        - exception_member_ids: List of member IDs to exclude from report
    
    Query Parameters (GET - legacy support):
        - month, year, account_id, team_id
    """
    try:
        # Handle both GET and POST requests
        if request.method == 'POST':
            data = request.get_json()
            month = data.get('month')
            year = data.get('year')
            account_id = data.get('account_id')
            team_id = data.get('team_id')
            rd_member_ids = data.get('rd_member_ids', [])
            rd_member_names = data.get('rd_member_names', [])
            exception_member_ids = data.get('exception_member_ids', [])
        else:
            # GET request - legacy support
            month = request.args.get('month', type=int)
            year = request.args.get('year', type=int)
            account_id = request.args.get('account_id', type=int)
            team_id = request.args.get('team_id', type=int)
            rd_member_ids = []
            rd_member_names = []
            exception_member_ids = []
        
        # Validate required parameters
        if not month or not year:
            now = datetime.now()
            month = month or now.month
            year = year or now.year
        
        # Determine account_id based on user role
        if current_user.role == 'super_admin':
            if not account_id:
                return jsonify({'error': 'Account ID is required for super admin'}), 400
        else:
            account_id = current_user.account_id
        
        # Validate team_id
        if not team_id:
            return jsonify({'error': 'Team ID is required'}), 400
        
        # Authorization check
        if current_user.role not in ['super_admin', 'account_admin', 'team_admin']:
            return jsonify({'error': 'Unauthorized - Admin access required'}), 403
        
        if current_user.role == 'account_admin' and current_user.account_id != account_id:
            return jsonify({'error': 'Unauthorized - Cannot access other accounts'}), 403
        
        if current_user.role == 'team_admin' and current_user.team_id != team_id:
            return jsonify({'error': 'Unauthorized - Cannot access other teams'}), 403
        
        # Generate the Excel file with custom RD and exception lists
        excel_buffer = generate_shift_allowance_excel(
            account_id=account_id,
            team_id=team_id,
            year=year,
            month=month,
            rd_member_ids=rd_member_ids,
            rd_member_names=rd_member_names,
            exception_member_ids=exception_member_ids
        )
        
        # Generate filename
        month_name = calendar.month_name[month]
        team = Team.query.get(team_id)
        team_name = team.name.replace(' ', '_') if team else 'Team'
        filename = f"Shift_Allowance_{team_name}_{month_name}_{year}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except ImportError as e:
        logger.error(f"Import error: {e}")
        return jsonify({'error': 'Excel generation library not available'}), 500
    except Exception as e:
        logger.error(f"Error generating shift allowance report: {e}")
        return jsonify({'error': str(e)}), 500


@shift_allowance_bp.route('/api/shift-allowance/preview', methods=['GET'])
@login_required
def preview_shift_allowance():
    """
    Preview shift allowance data as JSON (for debugging/testing).
    """
    try:
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        account_id = request.args.get('account_id', type=int)
        team_id = request.args.get('team_id', type=int)
        
        now = datetime.now()
        month = month or now.month
        year = year or now.year
        
        if current_user.role == 'super_admin':
            if not account_id:
                return jsonify({'error': 'Account ID required'}), 400
        else:
            account_id = current_user.account_id
        
        if not team_id:
            return jsonify({'error': 'Team ID required'}), 400
        
        # Get members
        members = TeamMember.query.filter_by(
            account_id=account_id,
            team_id=team_id,
            is_active=True
        ).all()
        
        rd_members = [m for m in members if is_rd_engineer(m.name)]
        team_members = [m for m in members if not is_rd_engineer(m.name)]
        
        return jsonify({
            'month': month,
            'year': year,
            'rd_count': len(rd_members),
            'team_count': len(team_members),
            'rd_members': [{'id': m.id, 'name': m.name, 'employee_id': m.employee_id} for m in rd_members],
            'team_members': [{'id': m.id, 'name': m.name, 'employee_id': m.employee_id} for m in team_members],
            'rd_engineer_names_config': RD_ENGINEER_NAMES
        })
        
    except Exception as e:
        logger.error(f"Error previewing shift allowance: {e}")
        return jsonify({'error': str(e)}), 500

